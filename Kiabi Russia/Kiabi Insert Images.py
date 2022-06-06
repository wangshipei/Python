import PIL
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import io
import urllib3
from urllib3.exceptions import LocationValueError
from tqdm import tqdm
from PIL import UnidentifiedImageError
import re


def insert_images(name, **urlplus):
    print(f'正在下载并插入图片的所有系列为Kiabi Russia-{name} ', '\n', urlplus.values())
    for a1 in urlplus.values():
        wb = load_workbook(f'Kiabi/Kiabi Russia-{name}-{a1}.xlsx')
        ws = wb.active

        for q1, i1 in tqdm(enumerate(ws['N'][1:], start=2), position=0, leave=True,
                           desc=f'正在下载并保存Kiabi Russia-{name}-{a1}的Pic-1'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i1.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'B{q1}')
            except LocationValueError:
                pass
            except PIL.UnidentifiedImageError:
                pass

        for q2, i2 in tqdm(enumerate(ws['O'][1:], start=2), position=0, leave=True,
                           desc=f'正在下载并保存Kiabi Russia-{name}-{a1}的Pic-2'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i2.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'I{q2}')
            except LocationValueError:
                pass
            except PIL.UnidentifiedImageError:
                pass

        for q3, i3 in tqdm(enumerate(ws['P'][1:], start=2), position=0, leave=True,
                           desc=f'正在下载并保存Kiabi Russia-{name}-{a1}的Pic-3'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i3.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'J{q3}')
            except LocationValueError:
                pass
            except PIL.UnidentifiedImageError:
                pass

        for q4, i4 in tqdm(enumerate(ws['Q'][1:], start=2), position=0, leave=True,
                           desc=f'正在下载并保存Kiabi Russia-{name}-{a1}的Pic-4'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i4.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'K{q4}')
            except LocationValueError:
                pass
            except PIL.UnidentifiedImageError:
                pass

        for q5, i5 in tqdm(enumerate(ws['R'][1:], start=2), position=0, leave=True,
                           desc=f'正在下载并保存Kiabi Russia-{name}-{a1}的Pic-5'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i5.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'L{q5}')
            except LocationValueError:
                pass
            except PIL.UnidentifiedImageError:
                pass

        print(f'Kiabi Russia-{name}-{a1}所有照片处理完毕!')
        wb.save(f'Kiabi/Kiabi Russia-{name}-{a1}.xlsx')


girlname = 'Girls'
girlurlplus = {
    # "platjya-yubki-devochki_254959": "dresses",
    # "verkhnyaya-odezhda-devochki_254835": "jackets",
    # "futbolki-vodolazki-devochki_254847": "T-shirts",
    # "rubashki-bluzki-devochki_254977": "shirts",
    # "dzhinsy-devochki_254891": "jeans",
    "bryuki-devochki_254857": "pants",
    "kombinezony-devochki_326072": "overall",
    "tolstovki-devochki_255007": "hoodies",
    "svitery-kardigany-devochki_254827": "sweaters",
    "sport-devochki_385552": "sports",
    "leginsy-devochki_255001": "leggings",
    "ukorochennye-bryuki-shorty-devochki_255031": "shorts",
    "pizhamy-khalaty-devochki_254939": "pajams",
    "nizhnee-belje-devochki_254949": "underwears",
    "chulochno-nosochnye-izdeliya-devochki_254971": "socks",
    "aksessuary-devochki_254983": "accessories"}
insert_images(girlname, **girlurlplus)
# 女童

boyname = 'Boys'
boyurlplus = {
    "futbolki-polo-malchiki_255435": "polos",
    "verkhnyaya-odezhda-malchiki_255423": "jackets",
    "dzhinsy-malchiki_255369": "jeans",
    "bryuki-malchiki_255409": "pants",
    "tolstovki-malchiki_255269": "hoodies",
    "svitery-kardigany-malchiki_255311": "sweaters",
    "rubashki-malchiki_255277": "shirts",
    "sport-malchiki_385577": "sports",
    "komplekty-malchiki_255287": "suits",
    "bermudy-shorty-malchiki_255393": "shorts",
    "ryukzaki-portfeli-malchiki_255361": "bags",
    "pizhamy-khalaty-malchiki_255337": "pajamas",
    "nizhnee-belje-malchiki_255317": "underwears",
    "noski-malchiki_255323": "socks",
    "aksessuary-malchiki_255295": "accessories"

}
insert_images(boyname, **boyurlplus)
# 男童

babyname = 'Baby'
babyurlplus = {"bodi-nizhnee-belje-malyshi_255739": "overall",
               "pizhamy-kombinezony-malyshi_255657": "sleepsuits",
               "konverty-dlya-novorozhdennykh-malyshi_255725": "sleepbags",
               "verkhnyaya-odezhda-malyshi_255717": "jackets",
               "futbolki-malyshi_255679": "T-shirts",
               "svitery-zhilety-tolstovki-malyshi_255691": "sweaters",
               "platjya-yubki-malyshi_255731": "dresses",
               "komplekty-bodi-pesochniki-malyshi_255705": "suits",
               "rubashki-malyshi_255699": "shirts",
               "bryuki-dzhinsy-leginsy-malyshi_255667": "pants",
               "shorty-bermudy-malyshi_255843": "shorts",
               "chulochno-nosochnye-izdeliya-kolgotki-malyshi_255751": "socks",
               "aksessuary-malyshi_255811": "accessories",
               "sport-malyshi_385598": "sports"}

insert_images(babyname, **babyurlplus)
# 婴儿
