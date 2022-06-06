from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import re
import glob
import os

print('正在加载文件......')
os.chdir(r'C:\Users\24910\PycharmProjects\SWOD\data')
file_list = glob.glob('*.xlsx')
CRED = '\033[91m'
CEND = '\033[0m'
for file in file_list:
    wb = load_workbook(file)
    ws = wb['СВОД']
    print('文件加载完成!')
    ERs = {'Cotton': 'Хлопок',
           'Polyester': 'полиэстер', 'Nylon': 'Нейлон', 'Viscose': 'Вискоза', 'Decoration': 'отделка',
           'Paper': 'Бумага', 'Tencel': 'Лиоцелл', 'Acrylic': 'Акрил', 'Elastane': 'Эластан', 'Linen': 'Лен',
           'Lurex': 'Люрекс', 'Wool': 'Шерсть',
           'Down': 'Пух', 'Feather': 'Перо', 'PVC coating': 'ПВХ-покрытие', 'PU': 'Полиуретан',
           'PU coating': 'полиуретан-покрытие', 'Body': 'корпус', 'Body lining': 'подкладка', 'Sleeve': 'рукава',
           'Sleeves lining': 'подкладка рукавов',
           'Hood': 'капюшон',
           'Hood lining': 'подкладка капюшона', 'Bottom': 'низ', 'Bottom lining': 'подкладка низа',
           'Combination': 'комбинированный', 'Top': 'верх',
           'Grey': 'серый', 'Mint': 'Мятный', 'Navy': 'Темно-синий', 'Shell': 'верх', 'Lining': 'подкладка',
           'Padding': 'утеплитель',
           'Pink': 'Розовый', 'Bule': 'синий', 'Fuchsia': 'фуксия', 'Dark Grey': 'темно-серый', 'Dark navy': 'темно-синий',
           'White': 'белый', 'Blue': 'голубой',
           'Beige': 'бежевый', 'Polyamide': 'Полиамид', 'Spandex': 'Спандекс', 'Lycra': 'Лайкра', 'Acetate': 'Ацетат',
           'Polyether': 'Полиэфир', 'Angora': 'Ангора',
           'Polypropylene': 'Полипропиллен', 'Polyacryl': 'Полиакрил', 'Polyurethane': 'Полиуретан',
           'Metallized fiber': 'Металлизированная нить',
           'Silk': 'Шелк', 'Straw': 'Солома', 'Lyocell': 'Лиоцелл', 'Natural Leather': 'Натуральная кожа',
           'Pvc': 'поливинилхлорид (полимерный/nматериал)',
           'Waistband': 'пояс', 'Main': 'основной', 'Collar': 'воротник', 'Composite': 'композитный',
           'Front lining': 'Подкладка передней части',
           'Combined': 'комбинированный', 'Facing': 'отделка', 'Front side': 'передняя часть',
           'Flip side': 'оборотная сторона', 'PA coating': 'полиамид-покрытие'}

    fill = PatternFill(start_color='00FF0033', fill_type='solid')
    print('正在比对面料/里布/含棉的英文和俄文信息......')
    # 面料
    for s in range(10, ws.max_row + 1):

        if ws.cell(row=s, column=30).value is None or ws.cell(row=s, column=30).value == 0:
            pass
        else:
            shellens = re.findall(r'(\d*%\w*)', ws.cell(row=s, column=30).value)
            shellrus = re.findall(r'(\d*%\w*)', ws.cell(row=s, column=33).value.replace(" ", ""))
            for shellen in shellens:
                shellen_lower = shellen.lower()
                for key1, value1 in ERs.items():
                    key1_lower = key1.lower()
                    value1_lower = value1.lower()
                    shellen_lower = shellen_lower.replace(key1_lower, value1_lower)
                if shellen_lower in shellrus:
                    pass
                else:
                    ws.cell(row=s, column=30).fill = fill
                    ws.cell(row=s, column=33).fill = fill
                    print(CRED + f"第{s}行，面料英文成分[{shellen}]对应俄文内容{shellrus}出错了！" + CEND)
    # 里布
    for ln in range(10, ws.max_row + 1):

        if ws.cell(row=ln, column=31).value is None or ws.cell(row=ln, column=31).value == 0:
            pass
        else:
            liningens = re.findall(r'(\d*%\w*)', ws.cell(row=ln, column=31).value)
            liningrus = re.findall(r'(\d*%\w*)', ws.cell(row=ln, column=34).value.replace(" ", ""))
            for liningen in liningens:
                liningen_lower = liningen.lower()
                for key2, value2 in ERs.items():
                    key2_lower = key2.lower()
                    value2_lower = value2.lower()
                    liningen_lower = liningen_lower.replace(key2_lower, value2_lower)
                if liningen_lower in liningrus:
                    pass
                else:
                    ws.cell(row=ln, column=31).fill = fill
                    ws.cell(row=ln, column=34).fill = fill
                    print(CRED + f"第{ln}行，面料英文成分[{liningen}]对应俄文内容{liningrus}出错了！" + CEND)

    # 含棉
    for pa in range(10, ws.max_row + 1):

        if ws.cell(row=pa, column=32).value is None or ws.cell(row=pa, column=32).value == 0:
            pass
        else:
            paddingens = re.findall(r'(\d*%\w*)', ws.cell(row=pa, column=32).value)
            paddingrus = re.findall(r'(\d*%\w*)', ws.cell(row=pa, column=35).value.replace(" ", ""))
            for paddingen in paddingens:
                paddingen_lower = paddingen.lower()
                for key3, value3 in ERs.items():
                    key3_lower = key3.lower()
                    value3_lower = value3.lower()
                    paddingen_lower = paddingen_lower.replace(key3_lower, value3_lower)
                if paddingen_lower in paddingrus:
                    pass
                else:
                    ws.cell(row=pa, column=32).fill = fill
                    ws.cell(row=pa, column=35).fill = fill
                    print(CRED + f"第{pa}行，面料英文成分[{paddingen}]对应俄文内容{paddingrus}出错了！" + CEND)
    print('面料/里布/含棉的英文和俄文信息所有数据比对完成!')

    # 下面开始计算百分比
    fill2 = PatternFill(start_color='00CC00FF', fill_type='solid')
    print('正在核对面料/里布/含棉百分比......')
    # 英文面料百分比
    for esn in range(10, ws.max_row + 1):
        if ws.cell(row=esn, column=30).value is None or ws.cell(row=esn, column=30).value == 0:
            pass
        else:
            en_shell_no1 = re.findall(r'(\d*)%\w*', ws.cell(row=esn, column=30).value)
            en_shell_no = list(map(int, en_shell_no1))
            en_shell_sum = sum(en_shell_no)
            if en_shell_sum == 100 or en_shell_sum == 200 or en_shell_sum == 300 or en_shell_sum == 0:
                pass
            else:
                ws.cell(row=esn, column=30).fill = fill2
                print(CRED + f"第{esn}行，英文面料百分比汇总不等于100%！" + CEND)
    # 俄文面料百分比
    for rsn in range(10, ws.max_row + 1):
        if ws.cell(row=rsn, column=33).value is None or ws.cell(row=rsn, column=33).value == 0:
            pass
        else:
            ru_shell_no1 = re.findall(r'(\d*)%\w*', ws.cell(row=rsn, column=33).value)
            ru_shell_no = list(map(int, ru_shell_no1))
            ru_shell_sum = sum(ru_shell_no)
            if ru_shell_sum == 100 or ru_shell_sum == 200 or ru_shell_sum == 300 or ru_shell_sum == 0:
                pass
            else:
                ws.cell(row=rsn, column=33).fill = fill2
                print(CRED + f"第{rsn}行，俄文面料百分比汇总不等于100%！" + CEND)

    # 英文里布百分比
    for eln in range(10, ws.max_row + 1):
        if ws.cell(row=eln, column=31).value is None or ws.cell(row=eln, column=31).value == 0:
            pass
        else:
            en_lining_no1 = re.findall(r'(\d*)%\w*', ws.cell(row=eln, column=31).value)
            en_lining_no = list(map(int, en_lining_no1))
            en_lining_sum = sum(en_lining_no)
            if en_lining_sum == 100 or en_lining_sum == 200 or en_lining_sum == 300 or en_lining_sum == 0:
                pass
            else:
                ws.cell(row=eln, column=31).fill = fill2
                print(CRED + f"第{eln}行，英文里布百分比汇总不等于100%！" + CEND)

    # 俄文里布百分比
    for rln in range(10, ws.max_row + 1):
        if ws.cell(row=rln, column=34).value is None or ws.cell(row=rln, column=34).value == 0:
            pass
        else:
            ru_lining_no1 = re.findall(r'(\d*)%\w*', ws.cell(row=rln, column=34).value)
            ru_lining_no = list(map(int, ru_lining_no1))
            ru_lining_sum = sum(ru_lining_no)
            if ru_lining_sum == 100 or ru_lining_sum == 200 or ru_lining_sum == 300 or ru_lining_sum == 0:
                pass
            else:
                ws.cell(row=rln, column=34).fill = fill2
                print(CRED + f"第{rln}行，俄文里布百分比汇总不等于100%！" + CEND)

    # 英文含棉百分比
    for epn in range(10, ws.max_row + 1):
        if ws.cell(row=epn, column=32).value is None or ws.cell(row=epn, column=32).value == 0:
            pass
        else:
            en_padding_no1 = re.findall(r'(\d*)%\w*', ws.cell(row=epn, column=32).value)
            en_padding_no = list(map(int, en_padding_no1))
            en_padding_sum = sum(en_padding_no)
            if en_padding_sum == 100 or en_padding_sum == 200 or en_padding_sum == 300 or en_padding_sum == 0:
                pass
            else:
                ws.cell(row=epn, column=32).fill = fill2
                print(CRED + f"第{epn}行，英文里布百分比汇总不等于100%！" + CEND)

    # 俄文含棉百分比
    for rpn in range(10, ws.max_row + 1):
        if ws.cell(row=rpn, column=35).value is None or ws.cell(row=rpn, column=35).value == 0:
            pass
        else:
            ru_padding_no1 = re.findall(r'(\d*)%\w*', ws.cell(row=rpn, column=35).value)
            ru_padding__no = list(map(int, ru_padding_no1))
            ru_padding__sum = sum(ru_padding__no)
            if ru_padding__sum == 100 or ru_padding__sum == 200 or ru_padding__sum == 300 or ru_padding__sum == 0:
                pass
            else:
                ws.cell(row=rpn, column=35).fill = fill2
                print(CRED + f"第{rpn}行，俄文里布百分比汇总不等于100%！" + CEND)
    print('面料/里布/含棉的百分比计算完成！!')
print('全部数据核对完成！')
wb.save(file)
