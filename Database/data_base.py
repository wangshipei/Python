from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import re
from tqdm import tqdm
import glob
import os

print('正在加载文件......')
os.chdir(r'C:\Users\24910\PycharmProjects\Database\data')
file_list = glob.glob('*.xlsx')
for file in file_list:
    wb = load_workbook(file)
    ws1 = wb['数据库-建宇']
    print('文件加载完成，正在处理数据......')
    season = {'HighSummer': 'HS', 'Spring': 'SP', 'Summer': 'SM', 'Autumn': 'AU', 'Winter': 'WI', 'School': 'SC',
              'Preautumn': 'PA', 'New Year': 'NY'}
    ERHs = {'Cotton': 'Cotton-Хлопок-Мақта', 'Polyester': 'Polyester-Полиэстер-Полиэстер',
            'Nylon': 'Nylon-Нейлон-Нейлон',
            'Viscose': 'Viscose-Вискоза-Вискоза', 'Decoration': 'Decoration-отделка-өрнек',
            'Paper': 'Paper-Бумага-Қағаз',
            'Tencel': 'Tencel-Лиоцелл-Лиоцелл', 'Acrylic': 'Acrylic-Акрил-Акрил',
            'Elastane': 'Elastane-Эластан-Эластан',
            'Linen': 'Linen-Лен-Зығыр', 'Lurex': 'Lurex-Люрекс-Люрекс', 'Wool': 'Wool-Шерсть-Жүн',
            'Down': 'Down-Пух-Мамық',
            'Feather': 'Feather-Перо-Қауырсын', 'PVC coating': 'PVC coating/покрытие-поливинилхлорид-поливинилхлорид',
            'PU': 'PU-Полиуретан-Полиуретан', 'PU coating': 'PU coating/покрытие-Полиуретан-Полиуретан',
            'Body': 'Body/корпус/дене', 'Body lining': 'Body lining-подкладка-астар', 'Sleeve': 'sleeve/рукава/жең',
            'Sleeves lining': 'Sleeves lining-подкладка рукавов-жеңдік астар', 'Hood': 'hood/капюшон/күләпара',
            'Hood lining': 'Hood lining-подкладка капюшона-күләпара астары', 'Bottom': 'Bottom-низ-етек',
            'Bottom lining': 'Bottom lining-подкладка низа-төменгі астар',
            'Combination': 'Combination/комбинированный /Құрамдастырылған', 'Top': 'Top/верх/тыс',
            'Grey': 'Grey/серый/сұр',
            'Mint': 'Mint/Мятный/жалбыз', 'Navy': 'Navy/Темно-синий/Қара-көк', 'Shell': 'Shell/верх/тыс',
            'Lining': 'Lining/подкладка/астар', 'Padding': 'Padding/утеплитель/жылытқыш',
            'Pink': 'Pink/Розовый/Қызғылт',
            'Bule': 'Bule-синий-көк', 'Fuchsia': 'Fuchsia-фуксия-фуксия',
            'Dark Grey': 'Dark Grey-темно-серый-күңгірт-сұр',
            'Dark navy': 'Dark navy-темно-синий-қара көк', 'White': 'White-белый-ақ', 'Blue': 'Blue-голубой-көгілдір',
            'Beige': 'Beige-бежевый-қоңыр-сарғыш', 'Polyamide': 'Polyamide-Полиамид-Полиамид',
            'Spandex': 'Spandex-Спандекс-Спандекс', 'Lycra': 'Lycra-Лайкра-Лайкра', 'Acetate': 'Acetate-Ацетат-Ацетат',
            'Polyether': 'Polyether-Полиэфир-Полиэфир', 'Angora': 'Angora-Ангора-Ангора',
            'Polypropylene': 'Polypropylene-Полипропиллен-Полипропилен', 'Polyacryl': 'Polyacryl-Полиакрил-Полиакрил',
            'Polyurethane': 'Polyurethane-Полиуретан-Полиуретан',
            'Metallized fiber': 'Metallized fiber-Металлизированная нить-Металдандырылған жіп',
            'Silk': 'Silk-Шелк-Жібек',
            'Straw': 'Straw-Солома-Сабан', 'Lyocell': 'Lyocell-Лиоцелл-Лиоцелл',
            'Natural Leather': 'Natural Leather-Натуральная кожа-Табиғи былғары',
            'Pvc': 'PVC-поливинилхлорид (полимерный/nматериал)-поливинилхлорид (полимерлі/nматериал)',
            'Waistband': 'Waistband-пояс-белдік', 'Main': 'main-основной-негізгі', 'Collar': 'collar-воротник-жаға',
            'Composite': 'composite-композитный-композиттік',
            'Front lining': 'Front lining-Подкладка передней части-Алдыңғы бөліктің астары',
            'Combined': 'Combined-комбинированный-құрамдастырылған', 'Facing': 'Facing-отделка-өрнек',
            'Front side': 'Front side-передняя часть-алдыңғы бөлік',
            'Flip side': 'Flip side-оборотная сторона-сыртқы жақ',
            'PA coating': 'PA coating/покрытие: Полиакрил/жабу: Полиакрил'}
    seasonpttn = re.compile(r'(.*)\d{4}')
    font = Font(name='Arial', size='10')
    border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True)
    fill = PatternFill(start_color='00C0C0C0', fill_type='solid')

    for s1 in tqdm(range(2, ws1.max_row + 1), position=0, leave=True, desc=f'正在处理第一张表单<数据库-建宇>'):
        for j1 in range(1, ws1.max_column + 1):
            ws1.cell(row=s1, column=j1).font = font
            ws1.cell(row=s1, column=j1).border = border
            ws1.cell(row=s1, column=j1).alignment = alignment

            ws1.cell(row=s1, column=5).fill = fill
            ws1.cell(row=s1, column=6).fill = fill
            ws1.cell(row=s1, column=7).fill = fill
            ws1.cell(row=s1, column=12).fill = fill
            ws1.cell(row=s1, column=27).fill = fill
            ws1.cell(row=s1, column=29).fill = fill

        ws1.cell(row=s1, column=5).value = ''.join(re.findall(r'.*(\d{4})', ws1.cell(row=s1, column=4).value))  # 数据库-年度
        shortseason = ''.join(re.findall(seasonpttn, ws1.cell(row=s1, column=4).value))
        ws1.cell(row=s1, column=6).value = season[shortseason.replace(' ', '')] + ws1.cell(row=s1,
                                                                                           column=5).value  # 数据库-中包季节
        ws1.cell(row=s1, column=7).value = ws1.cell(row=s1, column=5).value + '.' + season[
            shortseason.replace(' ', '')]  # 数据库-外箱季节
        ws1.cell(row=s1, column=27).value = ws1.cell(row=s1, column=26).value / ws1.cell(row=s1,
                                                                                         column=25).value  # 数据库-B1+乌拉尔配比包数
        ws1.cell(row=s1, column=29).value = ws1.cell(row=s1, column=28).value / ws1.cell(row=s1,
                                                                                         column=25).value  # 数据库-网店配比包数

        if ws1.cell(row=s1, column=9).value == 'Girls 2-6' or ws1.cell(row=s1, column=9).value == 'Girls 7-12':
            ws1.cell(row=s1, column=12).value = '女童'
        elif ws1.cell(row=s1, column=9).value == 'Boys 2-6' or ws1.cell(row=s1, column=9).value == 'Boys 7-12':
            ws1.cell(row=s1, column=12).value = '男童'
        elif ws1.cell(row=s1, column=9).value == 'Girls 0-24' or ws1.cell(row=s1, column=9).value == 'Boys 0-24':
            ws1.cell(row=s1, column=12).value = '婴儿'
        else:
            pass

    # print('第一张表单<数据库-建宇>处理完成！')

    ws2 = wb['洗水标-建宇']
    for s2 in tqdm(range(2, ws1.max_row + 1), position=0, leave=True, desc=f'正在处理第二张表单<洗水标-建宇>'):
        for j2 in range(1, ws2.max_column + 1):
            ws2.cell(row=s2, column=j2).font = font
            ws2.cell(row=s2, column=j2).border = border
            ws2.cell(row=s2, column=j2).alignment = alignment
            ws2.cell(row=s2, column=j2).fill = fill

        ws2.cell(row=s2, column=1).value = ws1.cell(row=s2, column=1).value  # 洗水标-工厂
        ws2.cell(row=s2, column=2).value = ws1.cell(row=s2, column=3).value  # 洗水标-款号
        shell = ws1.cell(row=s2, column=13).value
        lining = ws1.cell(row=s2, column=14).value
        padding = ws1.cell(row=s2, column=15).value
        for key1, value1 in ERHs.items():
            if ws1.cell(row=s2, column=13).value is None or ws1.cell(row=s2, column=13).value == 0:
                pass
            else:
                shell = shell.replace(key1, value1)
        for key2, value2 in ERHs.items():
            if ws1.cell(row=s2, column=14).value is None or ws1.cell(row=s2, column=14).value == 0:
                pass
            else:
                lining = lining.replace(key2, value2)
        for key3, value3 in ERHs.items():
            if ws1.cell(row=s2, column=15).value is None or ws1.cell(row=s2, column=15).value == 0:
                pass
            else:
                padding = padding.replace(key3, value3)

        ws2.cell(row=s2, column=3).value = shell  # 洗水标-面布
        ws2.cell(row=s2, column=4).value = lining  # 洗水标-里布
        ws2.cell(row=s2, column=5).value = padding  # 洗水标-含棉

        if ws2.cell(row=s2, column=5).value is None or ws2.cell(row=s2, column=5).value == 0 and ws2.cell(row=s2,
                                                                                                          column=4).value is None or ws2.cell(
                row=s2, column=4).value == 0:
            ws2.cell(row=s2, column=6).value = ws2.cell(row=s2, column=3).value
        elif ws2.cell(row=s2, column=5).value is not None and ws2.cell(row=s2, column=5).value != 0:
            ws2.cell(row=s2,
                     column=6).value = f'Shell/верх/тыс:\n{ws2.cell(row=s2, column=3).value}\nLining/подкладка/астар:\n{ws2.cell(row=s2, column=4).value}\nPadding/утеплитель/жылытқыш:\n{ws2.cell(row=s2, column=5).value}'
        else:
            ws2.cell(row=s2,
                     column=6).value = f'Shell/верх/тыс:\n{ws2.cell(row=s2, column=3).value}\nLining/подкладка/астар:\n{ws2.cell(row=s2, column=4).value}'

        ws2.cell(row=s2, column=7).value = ws1.cell(row=s2, column=19).value  # 洗水标-尺码
        ws2.cell(row=s2, column=8).value = ws1.cell(row=s2, column=20).value  # 洗水标-洗水标条形码
        ws2.cell(row=s2, column=9).value = ws1.cell(row=s2, column=42).value  # 洗水标-模版
    # print('第二张表单<洗水标-建宇>处理完成！')

    ws3 = wb['B1清单-建宇']
    for s3 in tqdm(range(2, ws1.max_row + 1), position=0, leave=True, desc=f'正在处理第三张表单<B1清单-建宇>'):
        for j3 in range(1, ws3.max_column + 1):
            ws3.cell(row=s3, column=j3).font = font
            ws3.cell(row=s3, column=j3).border = border
            ws3.cell(row=s3, column=j3).alignment = alignment
            ws3.cell(row=s3, column=j3).fill = fill
        ws3.cell(row=s3, column=1).value = ws1.cell(row=s3, column=1).value  # B1清单-工厂
        ws3.cell(row=s3, column=2).value = ws1.cell(row=s3, column=3).value  # B1清单-款号
        ws3.cell(row=s3, column=4).value = ws1.cell(row=s3, column=21).value  # B1清单-内箱条形码
        ws3.cell(row=s3, column=5).value = ws1.cell(row=s3, column=16).value  # B1清单-颜色
        ws3.cell(row=s3, column=6).value = ws1.cell(row=s3, column=19).value  # B1清单-尺码
        ws3.cell(row=s3, column=7).value = ws1.cell(row=s3, column=25).value  # B1清单-尺码配比
        ws3.cell(row=s3, column=8).value = ws1.cell(row=s3, column=20).value  # B1清单-洗水标条形码
        ws3.cell(row=s3, column=20).value = ws1.cell(row=s3, column=12).value  # B1清单-男女婴儿
    # print('第三张表单<B1清单-建宇>处理完成！')

    ws4 = wb['乌拉尔清单-建宇']
    for s4 in tqdm(range(2, ws1.max_row + 1), position=0, leave=True, desc=f'正在处理第四张表单<乌拉尔清单-建宇>'):
        for j4 in range(1, ws4.max_column + 1):
            ws4.cell(row=s4, column=j4).font = font
            ws4.cell(row=s4, column=j4).border = border
            ws4.cell(row=s4, column=j4).alignment = alignment
            ws4.cell(row=s4, column=j4).fill = fill
        ws4.cell(row=s4, column=1).value = ws1.cell(row=s4, column=1).value  # 乌拉尔清单-工厂
        ws4.cell(row=s4, column=2).value = ws1.cell(row=s4, column=3).value  # 乌拉尔清单-款号
        ws4.cell(row=s4, column=4).value = ws1.cell(row=s4, column=21).value  # 乌拉尔清单-内箱条形码
        ws4.cell(row=s4, column=5).value = ws1.cell(row=s4, column=16).value  # 乌拉尔清单-颜色
        ws4.cell(row=s4, column=6).value = ws1.cell(row=s4, column=19).value  # 乌拉尔清单-尺码
        ws4.cell(row=s4, column=7).value = ws1.cell(row=s4, column=25).value  # 乌拉尔清单-尺码配比
        ws4.cell(row=s4, column=8).value = ws1.cell(row=s4, column=20).value  # 乌拉尔清单-洗水标条形码
        ws4.cell(row=s4, column=20).value = ws1.cell(row=s4, column=12).value  # 乌拉尔清单-男女婴儿
    # print('第四张表单<乌拉尔清单-建宇>处理完成！')

    ws5 = wb['网店清单-建宇']
    for s5 in tqdm(range(2, ws1.max_row + 1), position=0, leave=True, desc=f'正在处理第五张表单<网店清单-建宇>'):
        for j5 in range(1, ws5.max_column + 1):
            ws5.cell(row=s5, column=j5).font = font
            ws5.cell(row=s5, column=j5).border = border
            ws5.cell(row=s5, column=j5).alignment = alignment
            ws5.cell(row=s5, column=j5).fill = fill
        ws5.cell(row=s5, column=1).value = ws1.cell(row=s5, column=1).value  # 网店清单-工厂
        ws5.cell(row=s5, column=2).value = ws1.cell(row=s5, column=3).value  # 网店清单-款号
        ws5.cell(row=s5, column=4).value = ws1.cell(row=s5, column=21).value  # 网店清单-内箱条形码
        ws5.cell(row=s5, column=5).value = ws1.cell(row=s5, column=16).value  # 网店清单-颜色
        ws5.cell(row=s5, column=6).value = ws1.cell(row=s5, column=19).value  # 网店清单-尺码
        ws5.cell(row=s5, column=7).value = ws1.cell(row=s5, column=25).value  # 网店清单-尺码配比
        ws5.cell(row=s5, column=8).value = ws1.cell(row=s5, column=20).value  # 网店清单-洗水标条形码
        ws5.cell(row=s5, column=20).value = ws1.cell(row=s5, column=12).value  # 网店清单-男女婴儿
    # print('第五张表单<网店清单-建宇>处理完成！')

    ws6 = wb['中包袋贴纸-辅料厂']
    for s6 in tqdm(range(2, ws1.max_row + 1), position=0, leave=True, desc=f'正在处理第六张表单<中包袋贴纸-辅料厂>'):
        for j6 in range(1, ws6.max_column + 1):
            ws6.cell(row=s6, column=j6).font = font
            ws6.cell(row=s6, column=j6).border = border
            ws6.cell(row=s6, column=j6).alignment = alignment
            ws6.cell(row=s6, column=j6).fill = fill
        ws6.cell(row=s6, column=1).value = ws1.cell(row=s6, column=1).value  # 中包袋贴纸-工厂
        ws6.cell(row=s6, column=2).value = ws1.cell(row=s6, column=3).value  # 中包袋贴纸-款号
        ws6.cell(row=s6, column=3).value = ws1.cell(row=s6, column=6).value  # 中包袋贴纸-中包季节
        ws6.cell(row=s6, column=4).value = ws1.cell(row=s6, column=21).value  # 中包袋贴纸-B1內箱条码
        ws6.cell(row=s6, column=5).value = ws1.cell(row=s6, column=22).value  # 中包袋贴纸-网店內箱条码
        ws6.cell(row=s6, column=6).value = ws1.cell(row=s6, column=16).value  # 中包袋贴纸-颜色
        ws6.cell(row=s6, column=7).value = ws1.cell(row=s6, column=19).value  # 中包袋贴纸-尺码
        ws6.cell(row=s6, column=8).value = ws1.cell(row=s6, column=25).value  # 中包袋贴纸-尺码配比
        ws6.cell(row=s6, column=9).value = ws1.cell(row=s6, column=20).value  # 中包袋贴纸-洗水标条码
        ws6.cell(row=s6, column=11).value = ws1.cell(row=s6, column=27).value  # 中包袋贴纸-B1配比包数
        ws6.cell(row=s6, column=12).value = ws1.cell(row=s6, column=29).value  # 中包袋贴纸-网店配比包数
        ws6.cell(row=s6, column=10).value = ws6.cell(row=s6, column=11).value + ws6.cell(row=s6,
                                                                                         column=12).value  # 中包袋贴纸-总配比包数
        ws6.cell(row=s6, column=13).value = ws1.cell(row=s6, column=40).value  # 中包袋贴纸-是否有二维码
    # print('第六张表单<中包袋贴纸-辅料厂>处理完成！')

    ws0 = wb['装箱建议-工厂']

    ws7 = wb['B1外箱贴纸-辅料厂']
    for s7 in tqdm(range(3, ws0.max_row + 1), position=0, leave=True, desc=f'正在处理第七张表单<B1外箱贴纸-辅料厂>'):
        for j7 in range(1, ws7.max_column + 1):
            ws7.cell(row=s7 - 1, column=j7).font = font
            ws7.cell(row=s7 - 1, column=j7).border = border
            ws7.cell(row=s7 - 1, column=j7).alignment = alignment
            ws7.cell(row=s7 - 1, column=j7).fill = fill

        ws7.cell(row=s7 - 1, column=1).value = ws0.cell(row=s7, column=1).value  # B1外箱贴纸-工厂
        ws7.cell(row=s7 - 1, column=2).value = ws0.cell(row=s7, column=5).value  # B1外箱贴纸-款号

        for k1 in range(2, ws1.max_row + 1):
            if ws7.cell(row=s7 - 1, column=2).value == ws1.cell(row=k1, column=3).value:
                ws7.cell(row=s7 - 1, column=3).value = ws1.cell(row=k1, column=7).value  # B1外箱贴纸-外箱季节
                ws7.cell(row=s7 - 1, column=4).value = ws1.cell(row=k1, column=21).value  # B1外箱贴纸-内箱中包袋条形码内容
                ws7.cell(row=s7 - 1, column=8).value = ws1.cell(row=k1, column=40).value  # B1外箱贴纸-是否有二维码
                ws7.cell(row=s7 - 1, column=17).value = ws1.cell(row=k1, column=12).value  # B1外箱贴纸-男/女/婴儿
    # print('第七张表单<B1外箱贴纸-辅料厂>处理完成！')

    ws8 = wb['乌拉尔外箱贴纸-辅料厂']
    for s8 in tqdm(range(3, ws0.max_row + 1), position=0, leave=True, desc=f'正在处理第八张表单<乌拉尔外箱贴纸-辅料厂>'):
        for j8 in range(1, ws8.max_column + 1):
            ws8.cell(row=s8 - 1, column=j8).font = font
            ws8.cell(row=s8 - 1, column=j8).border = border
            ws8.cell(row=s8 - 1, column=j8).alignment = alignment
            ws8.cell(row=s8 - 1, column=j8).fill = fill
        ws8.cell(row=s8 - 1, column=1).value = ws0.cell(row=s8, column=1).value  # 乌拉尔外箱贴纸-工厂
        ws8.cell(row=s8 - 1, column=2).value = ws0.cell(row=s8, column=5).value  # 乌拉尔外箱贴纸-款号

        for k2 in range(2, ws1.max_row + 1):
            if ws8.cell(row=s8 - 1, column=2).value == ws1.cell(row=k2, column=3).value:
                ws8.cell(row=s8 - 1, column=3).value = ws1.cell(row=k2, column=7).value  # 乌拉尔外箱贴纸-外箱季节
                ws8.cell(row=s8 - 1, column=4).value = ws1.cell(row=k2, column=21).value  # 乌拉尔外箱贴纸-内箱中包袋条形码内容
                ws8.cell(row=s8 - 1, column=8).value = ws1.cell(row=k2, column=40).value  # 乌拉尔外箱贴纸-是否有二维码
                ws8.cell(row=s8 - 1, column=17).value = ws1.cell(row=k2, column=12).value  # 乌拉尔外箱贴纸-男/女/婴儿
    # print('第八张表单<乌拉尔外箱贴纸-辅料厂>处理完成！')

    ws9 = wb['网店外箱贴纸-辅料厂']
    for s9 in tqdm(range(3, ws1.max_row + 1), position=0, leave=True, desc=f'正在处理第九张表单<网店外箱贴纸-辅料厂>'):
        for j9 in range(1, ws9.max_column + 1):
            ws9.cell(row=s9 - 1, column=j9).font = font
            ws9.cell(row=s9 - 1, column=j9).border = border
            ws9.cell(row=s9 - 1, column=j9).alignment = alignment
            ws9.cell(row=s9 - 1, column=j9).fill = fill
        ws9.cell(row=s9, column=1).value = ws1.cell(row=s9, column=1).value  # 网店外箱贴纸-工厂
        ws9.cell(row=s9, column=2).value = ws1.cell(row=s9, column=3).value  # 网店外箱贴纸-款号

        for k3 in range(2, ws1.max_row + 1):
            if ws9.cell(row=s9 - 1, column=2).value == ws1.cell(row=k3, column=3).value:
                ws9.cell(row=s9 - 1, column=3).value = ws1.cell(row=k3, column=7).value  # 网店外箱贴纸-外箱季节
                ws9.cell(row=s9 - 1, column=4).value = ws1.cell(row=k3, column=22).value  # 网店外箱贴纸-内箱中包袋条形码内容
                ws9.cell(row=s9 - 1, column=8).value = ws1.cell(row=k3, column=40).value  # 网店外箱贴纸-是否有二维码
                ws9.cell(row=s9 - 1, column=17).value = ws1.cell(row=k3, column=12).value  # 网店外箱贴纸-男/女/婴儿
    # print('第九张表单<网店外箱贴纸-辅料厂>处理完成！')

    ws10 = wb['辅料总表-辅料厂 工厂']
    for s10 in tqdm(range(2, ws1.max_row + 1), position=0, leave=True, desc=f'正在处理第十张表单<辅料总表-辅料厂 工厂>'):
        for j10 in range(1, 34):
            ws10.cell(row=s10, column=j10).font = font
            ws10.cell(row=s10, column=j10).border = border
            ws10.cell(row=s10, column=j10).alignment = alignment
            ws10.cell(row=s10, column=j10).fill = fill
        ws10.cell(row=s10, column=1).value = ws1.cell(row=s10, column=1).value  # 辅料总表-工厂
        ws10.cell(row=s10, column=3).value = ws1.cell(row=s10, column=3).value  # 辅料总表-款号
        ws10.cell(row=s10, column=4).value = ws1.cell(row=s10, column=8).value  # 辅料总表-品牌
        ws10.cell(row=s10, column=5).value = ws1.cell(row=s10, column=9).value  # 辅料总表-码段
        ws10.cell(row=s10, column=6).value = ws1.cell(row=s10, column=4).value  # 辅料总表-季节
        ws10.cell(row=s10, column=7).value = ws1.cell(row=s10, column=10).value  # 辅料总表-品名
        ws10.cell(row=s10, column=8).value = ws1.cell(row=s10, column=12).value  # 辅料总表-男/女
        ws10.cell(row=s10, column=9).value = ws1.cell(row=s10, column=13).value  # 辅料总表-面布
        ws10.cell(row=s10, column=10).value = ws1.cell(row=s10, column=14).value  # 辅料总表-里布
        ws10.cell(row=s10, column=11).value = ws1.cell(row=s10, column=15).value  # 辅料总表-含棉
        ws10.cell(row=s10, column=12).value = ws1.cell(row=s10, column=16).value  # 辅料总表-颜色（英文）
        ws10.cell(row=s10, column=13).value = ws1.cell(row=s10, column=17).value  # 辅料总表-颜色（俄文）
        ws10.cell(row=s10, column=14).value = ws1.cell(row=s10, column=18).value  # 辅料总表-颜色（中文）
        ws10.cell(row=s10, column=15).value = ws1.cell(row=s10, column=19).value  # 辅料总表-尺码
        ws10.cell(row=s10, column=16).value = ws1.cell(row=s10, column=23).value  # 辅料总表-套装件数
        ws10.cell(row=s10, column=17).value = ws1.cell(row=s10, column=20).value  # 辅料总表-洗水标条形码
        ws10.cell(row=s10, column=18).value = ws1.cell(row=s10, column=24).value  # 辅料总表-订单数量
        ws10.cell(row=s10, column=19).value = ws1.cell(row=s10, column=25).value  # 辅料总表-单配比包内件数
        ws10.cell(row=s10, column=20).value = ws1.cell(row=s10, column=29).value  # 辅料总表-单配比包内件数
        ws10.cell(row=s10, column=21).value = ws1.cell(row=s10, column=28).value  # 辅料总表-网店件数
        ws10.cell(row=s10, column=22).value = ws1.cell(row=s10, column=30).value  # 辅料总表-售价
        ws10.cell(row=s10, column=23).value = ws1.cell(row=s10, column=31).value  # 辅料总表-日期
        ws10.cell(row=s10, column=24).value = ws1.cell(row=s10, column=32).value  # 辅料总表-主标类型
        ws10.cell(row=s10, column=25).value = ws1.cell(row=s10, column=33).value  # 辅料总表-姓名标
        ws10.cell(row=s10, column=26).value = ws1.cell(row=s10, column=34).value  # 辅料总表-羊毛贴纸
        ws10.cell(row=s10, column=27).value = ws1.cell(row=s10, column=35).value  # 辅料总表-是否单胶袋
        ws10.cell(row=s10, column=28).value = ws1.cell(row=s10, column=36).value  # 辅料总表-防水吊牌
        ws10.cell(row=s10, column=29).value = ws1.cell(row=s10, column=37).value  # 辅料总表-触摸屏贴纸
        ws10.cell(row=s10, column=30).value = ws1.cell(row=s10, column=38).value  # 辅料总表-BCI吊牌
        ws10.cell(row=s10, column=31).value = ws1.cell(row=s10, column=39).value  # 辅料总表-GRS吊牌
        ws10.cell(row=s10, column=32).value = ws1.cell(row=s10, column=40).value  # 辅料总表-是否有二维码
        ws10.cell(row=s10, column=33).value = ws1.cell(row=s10, column=41).value  # 辅料总表-UPF吊牌
    # print('第十张表单<辅料总表-辅料厂 工厂>处理完成！')

    ws11 = wb['衣架总表-衣架厂']
    for s11 in tqdm(range(2, ws1.max_row + 1), position=0, leave=True, desc=f'正在处理第十一张表单<衣架总表-衣架厂>'):
        for j11 in range(1, 8):
            ws11.cell(row=s11, column=j11).font = font
            ws11.cell(row=s11, column=j11).border = border
            ws11.cell(row=s11, column=j11).alignment = alignment
            ws11.cell(row=s11, column=j11).fill = fill
        ws11.cell(row=s11, column=1).value = ws1.cell(row=s11, column=1).value  # 衣架总表-工厂
        ws11.cell(row=s11, column=3).value = ws1.cell(row=s11, column=3).value  # 衣架总表-款号
        ws11.cell(row=s11, column=4).value = ws1.cell(row=s11, column=8).value  # 衣架总表-品牌
        ws11.cell(row=s11, column=5).value = ws1.cell(row=s11, column=19).value  # 衣架总表-尺码
        ws11.cell(row=s11, column=6).value = ws1.cell(row=s11, column=9).value  # 衣架总表-尺码
        ws11.cell(row=s11, column=7).value = ws1.cell(row=s11, column=26).value  # 衣架总表-B1订单数量
    # print('第十一张表单<衣架总表-衣架厂>处理完成！')

    wb.save(file)
    print(f'{file}数据处理完成')
print('全部数据处理完成！')
