from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import io
import urllib3
from urllib3.exceptions import LocationValueError
from tqdm import tqdm
import re


def insert_images(name, *urlplus):
    a1 = [re.sub(r'\w*/\w*/', '', x1) for x1 in urlplus]
    print(f'正在下载并插入图片的所有系列为Reserved UK-{name}-{a1}')
    for a2 in a1:
        a3 = ''.join(a2)
        wb = load_workbook(f'Reserved/Reserved UK-{name}.xlsx')
        ws = wb.active

        for q1, i1 in tqdm(enumerate(ws['N'][1:], start=2), position=0, leave=True,
                           desc=f'正在下载并保存Reserved UK-{name}-{a3}的Pic-1'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i1.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'B{q1}')
            except LocationValueError:
                img = Image('1.jpg')
                ws.add_image(img, f'B{q1}')

        for q2, i2 in tqdm(enumerate(ws['O'][1:], start=2), position=0, leave=True,
                           desc=f'正在下载并保存Reserved UK-{name}-{a3}的Pic-2'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i2.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'I{q2}')
            except LocationValueError:
                img = Image('1.jpg')
                ws.add_image(img, f'I{q2}')

        for q3, i3 in tqdm(enumerate(ws['P'][1:], start=2), position=0, leave=True,
                           desc=f'正在下载并保存Reserved UK-{name}-{a3}的Pic-3'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i3.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'J{q3}')
            except LocationValueError:
                img = Image('1.jpg')
                ws.add_image(img, f'J{q3}')

        for q4, i4 in tqdm(enumerate(ws['Q'][1:], start=2), position=0, leave=True,
                           desc=f'正在下载并保存Reserved UK-{name}-{a3}的Pic-4'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i4.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'K{q4}')
            except LocationValueError:
                img = Image('1.jpg')
                ws.add_image(img, f'K{q4}')

        for q5, i5 in tqdm(enumerate(ws['R'][1:], start=2), position=0, leave=True,
                           desc=f'正在下载并保存Reserved UK-{name}-{a3}的Pic-5'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i5.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'L{q5}')
            except LocationValueError:
                img = Image('1.jpg')
                ws.add_image(img, f'L{q5}')

        print(f'Reserved UK-{name}-{a3}所有照片处理完毕!')
        wb.save(f'Reserved/Reserved UK-{name}.xlsx')


# girlname = 'Girl'
# girlurlplus = [
#     "girl/junior/longsleeve",
#     "girl/junior/t-shirts",
#     "girl/junior/shirts",
#     "girl/junior/sweatshirts",
#     "girl/junior/sweaters",
#     "girl/junior/trousers",
#     "girl/junior/jeans",
#     "girl/junior/skirts",
#     "girl/junior/sets",
#     "girl/junior/bags",
#     "girl/junior/garment"
# ]
# insert_images(girlname, *girlurlplus)

# boyname = 'Boy'
# boyurlplus = [
#     "boy/junior/sweatshirts",
#     "boy/junior/longsleeve",
#     "boy/junior/trousers",
#     "boy/junior/t-shirts",
#     "boy/junior/sweaters",
#     "boy/junior/shirts",
#     "boy/junior/bags",
#     "boy/junior/garment"
# ]
# insert_images(boyname, *boyurlplus)
# 男童

babyname = 'Baby'
babyurlplus = [
    # "girl/baby/sets",
    "boy/baby"
]
insert_images(babyname, *babyurlplus)
# 婴儿

newbornname = 'Newborn'
newbornurlplus = ["girl/newborn", "boy/newborn"]
insert_images(newbornname, *newbornurlplus)
# 新生儿
