from openpyxl import load_workbook
from tqdm import tqdm


def eshop_qrcode(EWarehouse, ART, PCS, EBOX, EMI, ELMI):
    wb = load_workbook(f'QR/IME_{ART}.xlsx')
    ws = wb.active
    ws.cell(row=3, column=5).value = '款号'
    ws.cell(row=3, column=6).value = '仓库和总箱号'
    ws.cell(row=3, column=7).value = '中包数目'
    Q = 0
    P = 1
    for b1 in tqdm(range(4, EBOX + 4), position=0, leave=True,
                   desc=f'正在处理{ART}的{EWarehouse}二维码'):
        M = 4 + EMI * PCS * Q
        N = 4 + EMI * PCS * P
        for b2 in range(M, N):
            ws.cell(row=b2, column=1).value = ws.cell(row=b1, column=8).value
            ws.cell(row=b2, column=5).value = ART
            ws.cell(row=b2, column=6).value = EWarehouse
            ws.cell(row=b2, column=7).value = EMI
        Q = Q + 1
        P = P + 1

    if ELMI != 0:
        for EL in range(0, ELMI * PCS):
            ws.cell(row=EBOX * EMI * PCS + EL + 4, column=1).value = ws.cell(row=EBOX + 4, column=8).value
            ws.cell(row=EBOX * EMI * PCS + EL + 4, column=5).value = ART
            ws.cell(row=EBOX * EMI * PCS + EL + 4, column=6).value = EWarehouse
            ws.cell(row=EBOX * EMI * PCS + EL + 4, column=7).value = ELMI

    ws2 = wb.copy_worksheet(ws)
    ws2.title = '简化版'
    wb.active = wb['简化版']
    q, p = 0, 0
    for i in tqdm(range(4, ws2.max_row + 1), position=0, leave=True,
                  desc=f'正在简化{ART}的{EWarehouse}中包袋条形码'):
        if ws2.cell(row=i + q + 1, column=2).value == ws2.cell(row=i + q, column=2).value:
            ws2.delete_rows(q + i + 1)
            q = q - 1
        else:
            pass
    for a in tqdm(range(4, ws2.max_row + 1), position=0, leave=True,
                  desc=f'正在简化{ART}的{EWarehouse}外箱条形码'):
        if ws2.cell(row=a + p, column=1).value is None:
            break
        elif ws2.cell(row=a + p + 1, column=1).value == ws2.cell(row=a + p, column=1).value:
            ws2.delete_rows(a + p + 1)
            p = p - 1
        else:
            pass
    ws.delete_cols(8)
    ws2.delete_cols(8)
    wb.save(f'QR/IME_{ART}.xlsx')
    print(f'{ART}-E-shop二维码数据处理完毕！')


def b1ural_qrcode(B1Warehouse, URALWarehouse, ART, PCS, B1BOX, B1MI, B1LMI, UBOX, UMI, ULMI):
    wb = load_workbook(f'QR/{ART}.xlsx')
    ws = wb.active
    ws.cell(row=3, column=5).value = '款号'
    ws.cell(row=3, column=6).value = '仓库和总箱号'
    ws.cell(row=3, column=7).value = '中包数目'
    Q1 = 0
    P1 = 1
    Q2 = 0
    P2 = 1
    for c1 in tqdm(range(4, B1BOX + 4), position=0, leave=True,
                   desc=f'正在处理{ART}的{B1Warehouse}二维码'):
        M1 = 4 + B1MI * PCS * Q1
        N1 = 4 + B1MI * PCS * P1
        for c2 in range(M1, N1):
            ws.cell(row=c2, column=1).value = ws.cell(row=c1, column=8).value
            ws.cell(row=c2, column=5).value = ART
            ws.cell(row=c2, column=6).value = B1Warehouse
            ws.cell(row=c2, column=7).value = B1MI
        Q1 = Q1 + 1
        P1 = P1 + 1

    if B1LMI != 0:
        for B1L in range(0, B1LMI * PCS):
            ws.cell(row=B1BOX * B1MI * PCS + B1L + 4, column=1).value = ws.cell(row=B1BOX + 4, column=8).value
            ws.cell(row=B1BOX * B1MI * PCS + B1L + 4, column=5).value = ART
            ws.cell(row=B1BOX * B1MI * PCS + B1L + 4, column=6).value = B1Warehouse
            ws.cell(row=B1BOX * B1MI * PCS + B1L + 4, column=7).value = B1LMI
    # URAL
    for d1 in tqdm(range(4, UBOX + 4), position=0, leave=True,
                   desc=f'正在处理{ART}的{URALWarehouse}二维码'):
        if B1LMI != 0:
            B1LBOX = 1
        else:
            B1LBOX = 0
        M2 = 4 + (B1BOX * B1MI + B1LBOX * B1LMI) * PCS + UMI * PCS * Q2
        N2 = 4 + (B1BOX * B1MI + B1LBOX * B1LMI) * PCS + UMI * PCS * P2
        for d2 in range(M2, N2):
            ws.cell(row=d2, column=1).value = ws.cell(row=d1, column=9).value
            ws.cell(row=d2, column=5).value = ART
            ws.cell(row=d2, column=6).value = URALWarehouse
            ws.cell(row=d2, column=7).value = B1MI
        Q2 = Q2 + 1
        P2 = P2 + 1

    if ULMI != 0:
        for UL in range(0, ULMI * PCS):
            if B1LMI != 0:
                B1LBOX = 1
            else:
                B1LBOX = 0
            ws.cell(row=(B1BOX * UMI + B1LBOX * ULMI) * PCS + UBOX * UMI * PCS + UL + 4,
                    column=1).value = ws.cell(row=UBOX + 4, column=9).value
            ws.cell(row=(B1BOX * UMI + B1LBOX * ULMI) * PCS + UBOX * UMI * PCS + UL + 4, column=5).value = ART
            ws.cell(row=(B1BOX * UMI + B1LBOX * ULMI) * PCS + UBOX * UMI * PCS + UL + 4,
                    column=6).value = URALWarehouse
            ws.cell(row=(B1BOX * UMI + B1LBOX * ULMI) * PCS + UBOX * UMI * PCS + UL + 4, column=7).value = ULMI

    ws2 = wb.copy_worksheet(ws)
    ws2.title = '简化版'
    wb.active = wb['简化版']
    q, p = 0, 0
    for i in tqdm(range(4, ws2.max_row + 1), position=0, leave=True,
                  desc=f'正在简化{ART}的{B1Warehouse}和{URALWarehouse}中包袋条形码'):
        if ws2.cell(row=i + q + 1, column=2).value == ws2.cell(row=i + q, column=2).value:
            ws2.delete_rows(q + i + 1)
            q = q - 1
        else:
            pass

    for a in tqdm(range(4, ws2.max_row + 1), position=0, leave=True,
                  desc=f'正在简化{ART}的{B1Warehouse}和{URALWarehouse}外箱条形码'):
        if ws2.cell(row=a + p, column=1).value is None:
            break
        elif ws2.cell(row=a + p + 1, column=1).value == ws2.cell(row=a + p, column=1).value:
            ws2.delete_rows(a + p + 1)
            p = p - 1
        else:
            pass
    ws.delete_cols(8, 2)
    ws2.delete_cols(8, 2)
    wb.save(f'QR/{ART}.xlsx')
    print(f'{ART}-B1&URAL二维码数据处理完毕！')


ART1 = 'W22FU3-G77-1tg-44'  # 款号
PCS1 = 11  # 中包内件数
EWarehouse1 = 'E-shop 总箱号376'  # 网店总箱号
EBOX1 = 186  # 网店箱数（不含尾箱）
EMI1 = 1  # 网店中包数
ELMI1 = 0  # 网店尾箱中包数
B1Warehouse1 = 'B1 总箱号123'  # B1总箱号
B1BOX1 = 123  # B1箱数（不含尾箱）
B1MI1 = 1  # B1中包数
B1LMI1 = 0  # B1尾箱中包数
URALWarehouse1 = 'URAL 总箱号640'  # 乌拉尔总箱
UBOX1 = 39  # URAL箱数（不含尾箱）
UMI1 = 1  # URAL中包数
ULMI1 = 0  # URAL尾箱中包数
# --------------第一个款式-----------------


# ART2 = 'W22FU3-G77-2tg-99'                  # 款号
# PCS2 = 11                                   # 中包内件数
# EWarehouse2 = 'E-shop 总箱号376'              # 网店总箱号
# EBOX2 = 190                                 # 网店箱数（不含尾箱）
# EMI2 = 1                                     # 网店中包数
# ELMI2 = 0                                    # 网店尾箱中包数
# B1Warehouse2 = 'B1 总箱号498'                 # B1总箱号
# B1BOX2 = 998                                # B1箱数（不含尾箱）
# B1MI2 = 1                                    # B1中包数
# B1LMI2 = 0                                   # B1尾箱中包数
# URALWarehouse2 = 'URAL 总箱号640'            # 乌拉尔总箱号
# UBOX2 = 266                                  # URAL箱数（不含尾箱）
# UMI2 = 1                                    # URAL中包数
# ULMI2 = 0                                    # URAL尾箱中包数
# # --------------第二个款式-----------------
#
#
# ART3 = 'W22FC5-B75kb-99'                                     # 款号
# PCS3 = 10                                       # 中包内件数
# EWarehouse3 = 'E-shop 总箱号400'             # 网店总箱号
# EBOX3 = 196                                      # 网店箱数（不含尾箱）
# EMI3 = 1                                        # 网店中包数
# ELMI3 = 0                                        # 网店尾箱中包数
# B1Warehouse3 = 'B1 总箱号712/318'                  # B1总箱号
# B1BOX3 = 1030                                        # B1箱数（不含尾箱）
# B1MI3 = 1                                         # B1中包数
# B1LMI3 = 0                                        # B1尾箱中包数
# URALWarehouse3 = 'URAL 总箱号699'                   # 乌拉尔总箱号
# UBOX3 = 274                                        # URAL箱数（不含尾箱）
# UMI3 = 1                                          # URAL中包数
# ULMI3 = 0                                        # URAL尾箱中包数
# # --------------第三个款式-----------------
#
#
# ART4 = 'W22FC5-B75tb-99'       # 款号
# PCS4 = 11            # 中包内件数
# EWarehouse4 = 'E-shop 总箱号593'           # 网店总箱号
# EBOX4 = 250          # 网店箱数（不含尾箱）
# EMI4 = 1           # 网店中包数
# ELMI4 = 0          # 网店尾箱中包数
# B1Warehouse4 = 'B1 总箱号586/586/818'              # B1总箱号
# B1BOX4 = 1310         # B1箱数（不含尾箱）
# B1MI4 = 1          # B1中包数
# B1LMI4 = 0         # B1尾箱中包数
# URALWarehouse4 = 'URAL 总箱号661'          # 乌拉尔总箱号
# UBOX4 = 349          # URAL箱数（不含尾箱）
# UMI4 = 1           # URAL中包数
# ULMI4 = 0          # URAL尾箱中包数
# # --------------第四个款式-----------------

ART5 = 'W22FU5-B133kb-88'       # 款号
PCS5 = 10           # 中包内件数
EWarehouse5 = 'E-shop 总箱号306'           # 网店总箱号
EBOX5 = 102           # 网店箱数（不含尾箱）
EMI5 = 2           # 网店中包数
ELMI5 = 0          # 网店尾箱中包数
B1Warehouse5 = 'B1 总箱号587'              # B1总箱号
B1BOX5 = 195         # B1箱数（不含尾箱）
B1MI5 = 2          # B1中包数
B1LMI5 = 1         # B1尾箱中包数
URALWarehouse5 = 'URAL 总箱号'          # 乌拉尔总箱号
UBOX5 = 52          # URAL箱数（不含尾箱）
UMI5 = 2           # URAL中包数
ULMI5 = 1          # URAL尾箱中包数
# --------------第五个款式-----------------
#
# ART6 = 'W22FU5-B133tb-66'       # 款号
# PCS6 = 10           # 中包内件数
# EWarehouse6 = 'E-shop 总箱号306'           # 网店总箱号
# EBOX6 = 204          # 网店箱数（不含尾箱）
# EMI6 = 1           # 网店中包数
# ELMI6 = 0          # 网店尾箱中包数
# B1Warehouse6 = 'B1 总箱号587'              # B1总箱号
# B1BOX6 = 391         # B1箱数（不含尾箱）
# B1MI6 = 1          # B1中包数
# B1LMI6 = 0         # B1尾箱中包数
# URALWarehouse6 = 'URAL 总箱号'          # 乌拉尔总箱号
# UBOX6 = 105          # URAL箱数（不含尾箱）
# UMI6 = 1           # URAL中包数
# ULMI6 = 0          # URAL尾箱中包数

# --------------第五个款式-----------------

eshop_qrcode(EWarehouse1, ART1, PCS1, EBOX1, EMI1, ELMI1)
# eshop_qrcode(EWarehouse2, ART2, PCS2, EBOX2, EMI2, ELMI2)
# eshop_qrcode(EWarehouse3, ART3, PCS3, EBOX3, EMI3, ELMI3)
# eshop_qrcode(EWarehouse4, ART4, PCS4, EBOX4, EMI4, ELMI4)
eshop_qrcode(EWarehouse5, ART5, PCS5, EBOX5, EMI5, ELMI5)
# eshop_qrcode(EWarehouse6, ART6, PCS6, EBOX6, EMI6, ELMI6)
b1ural_qrcode(B1Warehouse1, URALWarehouse1, ART1, PCS1, B1BOX1, B1MI1, B1LMI1, UBOX1, UMI1, ULMI1)
# b1ural_qrcode(B1Warehouse2, URALWarehouse2, ART2, PCS2, B1BOX2, B1MI2, B1LMI2, UBOX2, UMI2, ULMI2)
# b1ural_qrcode(B1Warehouse3, URALWarehouse3, ART3, PCS3, B1BOX3, B1MI3, B1LMI3, UBOX3, UMI3, ULMI3)
# b1ural_qrcode(B1Warehouse4, URALWarehouse4, ART4, PCS4, B1BOX4, B1MI4, B1LMI4, UBOX4, UMI4, ULMI4)
b1ural_qrcode(B1Warehouse5, URALWarehouse5, ART5, PCS5, B1BOX5, B1MI5, B1LMI5, UBOX5, UMI5, ULMI5)
# b1ural_qrcode(B1Warehouse6, URALWarehouse6, ART6, PCS6, B1BOX6, B1MI6, B1LMI6, UBOX6, UMI6, ULMI6)
