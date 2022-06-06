import requests
import base64
import glob
import os
from openpyxl import load_workbook
from tqdm import tqdm
import PyPDF2
from openpyxl.styles import PatternFill
from datetime import date

wb1 = load_workbook(r'C:\Users\黄洁尔\PycharmProjects\test\ED+INV+REG\增值税发票.xlsx')
ws1 = wb1.active

os.chdir(r'C:\Users\黄洁尔\PycharmProjects\test\ED+INV+REG\发票')
file_list = glob.glob('*.pdf')
mrow = ws1.max_row
today = date.today()
q = 1
for file in file_list:
    f1 = open(file, 'rb')
    readpdf = PyPDF2.PdfFileReader(f1)
    page = readpdf.numPages
    f = open(file, 'rb')
    f_name, f_ext = os.path.splitext(file)
    img = base64.b64encode(f.read())
    data = {}
    datas = []
    try:
        for p in tqdm(range(1, page + 1), position=0, leave=True, desc=f'正在识别发票{f_name}并写入数据'):
            q = q + 1
            params = {"pdf_file": img, 'pdf_file_num': p}
            request_url = "https://aip.baidubce.com/rest/2.0/ocr/v1/vat_invoice"
            access_token = '24.23225a9b00af2ec2cdda694294f08e53.2592000.1652316471.282335-25947484'
            request_url = request_url + "?access_token=" + access_token
            headers = {'content-type': 'application/x-www-form-urlencoded'}
            response = requests.post(request_url, data=params, headers=headers)
            if response:
                json1 = response.json()
                data['InvoiceNum'] = json1['words_result']['InvoiceNum']
                data['InvoiceDate'] = json1['words_result']['InvoiceDate']
                data['TotalAmount'] = json1['words_result']['TotalAmount']
                data['TotalTax'] = json1['words_result']['TotalTax']
                data['AmountInFiguers'] = json1['words_result']['AmountInFiguers']
                data['AmountInWords'] = json1['words_result']['AmountInWords']
                data['CommodityName'] = json1['words_result']['CommodityName'][0]['word']
                data['CommodityUnit'] = json1['words_result']['CommodityUnit'][0]['word']
                data['CommodityNum'] = json1['words_result']['CommodityNum'][0]['word']
                data['CommodityPrice'] = json1['words_result']['CommodityPrice'][0]['word']
                data['CommodityAmount'] = json1['words_result']['CommodityAmount'][0]['word']
                data['CommodityTaxRate'] = json1['words_result']['CommodityTaxRate'][0]['word']
                data['CommodityTax'] = json1['words_result']['CommodityTax'][0]['word']
                data['PurchaserName'] = json1['words_result']['PurchaserName']
                data['PurchaserRegisterNum'] = json1['words_result']['PurchaserRegisterNum']
                data['PurchaserAddress'] = json1['words_result']['PurchaserAddress']
                data['PurchaserBank'] = json1['words_result']['PurchaserBank']
                data['SellerName'] = json1['words_result']['SellerName']
                data['SellerRegisterNum'] = json1['words_result']['SellerRegisterNum']
                data['SellerAddress'] = json1['words_result']['SellerAddress']
                data['SellerBank'] = json1['words_result']['SellerBank']
                data['InvoiceTypeOrg'] = json1['words_result']['InvoiceTypeOrg']
                data['Province'] = json1['words_result']['Province']
                data['InvoiceType'] = json1['words_result']['InvoiceType']
                data['SheetNum'] = json1['words_result']['SheetNum']
                data['InvoiceCode'] = json1['words_result']['InvoiceCode']
                data['Password'] = json1['words_result']['Password']
                data['Payee'] = json1['words_result']['Payee']
                data['Checker'] = json1['words_result']['Checker']
                data['NoteDrawer'] = json1['words_result']['NoteDrawer']

                ws1.cell(row=mrow + q, column=1).value = today
                ws1.cell(row=mrow + q, column=2).value = f_name

                ws1.cell(row=mrow + q, column=3).value = data['InvoiceDate']
                ws1.cell(row=mrow + q, column=4).value = data['InvoiceNum']
                ws1.cell(row=mrow + q, column=5).value = data['TotalAmount']
                ws1.cell(row=mrow + q, column=6).value = data['TotalTax']
                ws1.cell(row=mrow + q, column=7).value = data['AmountInFiguers']
                ws1.cell(row=mrow + q, column=8).value = data['AmountInWords']
                ws1.cell(row=mrow + q, column=9).value = data['CommodityName']
                ws1.cell(row=mrow + q, column=10).value = data['CommodityUnit']
                ws1.cell(row=mrow + q, column=11).value = data['CommodityNum']
                ws1.cell(row=mrow + q, column=12).value = data['CommodityPrice']
                ws1.cell(row=mrow + q, column=13).value = data['CommodityAmount']
                ws1.cell(row=mrow + q, column=14).value = data['CommodityTaxRate']
                ws1.cell(row=mrow + q, column=15).value = data['CommodityTax']
                ws1.cell(row=mrow + q, column=16).value = data['PurchaserName']
                ws1.cell(row=mrow + q, column=17).value = data['PurchaserRegisterNum']
                ws1.cell(row=mrow + q, column=18).value = data['PurchaserAddress']
                ws1.cell(row=mrow + q, column=19).value = data['PurchaserBank']
                ws1.cell(row=mrow + q, column=20).value = data['SellerName']
                ws1.cell(row=mrow + q, column=21).value = data['SellerRegisterNum']
                ws1.cell(row=mrow + q, column=22).value = data['SellerAddress']
                ws1.cell(row=mrow + q, column=23).value = data['SellerBank']
                ws1.cell(row=mrow + q, column=24).value = data['InvoiceTypeOrg']
                ws1.cell(row=mrow + q, column=25).value = data['Province']
                ws1.cell(row=mrow + q, column=26).value = data['InvoiceType']
                ws1.cell(row=mrow + q, column=27).value = data['SheetNum']
                ws1.cell(row=mrow + q, column=28).value = data['InvoiceCode']
                ws1.cell(row=mrow + q, column=29).value = data['Password']
                ws1.cell(row=mrow + q, column=30).value = data['Payee']
                ws1.cell(row=mrow + q, column=31).value = data['Checker']
                ws1.cell(row=mrow + q, column=32).value = data['NoteDrawer']
                if float(ws1.cell(row=mrow + q, column=5).value) + float(ws1.cell(row=mrow + q, column=6).value) != float(
                        ws1.cell(row=mrow + q, column=7).value):
                    ws1.cell(row=mrow + q, column=7).fill = PatternFill(start_color='FFC7CE', fill_type='solid')
    except Exception as e:
        print(f'错误信息：{e}')
        pass
wb1.save(r'C:\Users\黄洁尔\PycharmProjects\test\ED+INV+REG\增值税发票.xlsx')

wb2 = load_workbook(r'C:\Users\黄洁尔\PycharmProjects\test\ED+INV+REG\增值税发票.xlsx')
ws2 = wb2.active

wb3 = load_workbook(r'C:\Users\黄洁尔\PycharmProjects\test\ED+INV+REG\开票资料.xlsx')
ws3 = wb3.active

for i in tqdm(range(2, ws3.max_row + 1), position=0, leave=True, desc=f'正在写入发票数据'):
    for j in range(2, ws2.max_row + 1):
        if ws3.cell(row=i, column=2).value is None:
            pass
        elif ws2.cell(row=j, column=2).value is not None and f'{ws3.cell(row=i, column=2).value}({ws3.cell(row=i, column=3).value})' == ws2.cell(row=j, column=2).value.replace(' ', ''):
            ws3.cell(row=i, column=14).value = ws2.cell(row=j, column=3).value
            ws3.cell(row=i, column=15).value = ws2.cell(row=j, column=4).value
            ws3.cell(row=i, column=16).value = ws2.cell(row=j, column=5).value
            ws3.cell(row=i, column=17).value = ws2.cell(row=j, column=6).value
            ws3.cell(row=i, column=18).value = ws2.cell(row=j, column=7).value
        else:
            pass
wb3.save(r'C:\Users\黄洁尔\PycharmProjects\test\ED+INV+REG\开票资料.xlsx')
