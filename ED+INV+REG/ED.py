import glob
import os
import re
from openpyxl import load_workbook
from tqdm import tqdm

os.chdir(r'C:\Users\黄洁尔\PycharmProjects\test\ED+INV+REG\关单')
file_list = glob.glob('报关资料*.xlsx')
wb1 = load_workbook(r'C:\Users\黄洁尔\PycharmProjects\test\ED+INV+REG\开票资料.xlsx')
ws1 = wb1.active
mrow = ws1.max_row
q = 0
for file in tqdm(file_list, position=0, leave=True,
                 desc=f'正在识别关单并写入数据'):
    f_name, f_ext = os.path.splitext(file)
    factory = ''.join(re.findall(r'报关资料 QH.* (.*) \d*.*CTNS', f_name))
    wb2 = load_workbook(file, data_only=True)
    ws2 = wb2['要素']
    ws22 = wb2['单位']

    with_time = str(ws2.cell(row=11, column=2).value)
    without_time = with_time.replace('00:00:00', '')

    for i in range(25, ws2.max_row + 1):
        if ws2.cell(row=i, column=4).value is None:
            break
        else:
            ws1.cell(row=mrow + i - 23 + q, column=3).value = ws2.cell(row=i, column=1).value  # 关单项号
            ws1.cell(row=mrow + i - 23 + q, column=10).value = ws2.cell(row=i, column=2).value  # 款号
            ws1.cell(row=mrow + i - 23 + q, column=5).value = ws2.cell(row=i, column=4).value  # 品名
            ws1.cell(row=mrow + i - 23 + q, column=13).value = ws2.cell(row=i, column=10).value  # 报关外金额
            ws1.cell(row=mrow + i - 23 + q, column=6).value = ws2.cell(row=i, column=11).value  # 数量
            ws1.cell(row=mrow + i - 23 + q, column=1).value = without_time  # 日期
            ws1.cell(row=mrow + i - 23 + q, column=2).value = ws2.cell(row=8, column=2).value  # 关单号
            ws1.cell(row=mrow + i - 23 + q, column=4).value = factory  # 工厂
            for index, d in enumerate(ws22['B'], start=1):
                if d.value == ws2.cell(row=i, column=4).value:
                    ws1.cell(row=mrow + i - 23 + q, column=7).value = ws22.cell(row=index, column=4).value  # 单位
    q = q + 2
    wb1.save(r'C:\Users\黄洁尔\PycharmProjects\test\ED+INV+REG\开票资料.xlsx')
    wb2.close()
