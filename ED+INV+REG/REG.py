from openpyxl import load_workbook
from tqdm import tqdm
import shutil
from datetime import date

wb = load_workbook(r'C:\Users\黄洁尔\PycharmProjects\test\ED+INV+REG\开票资料.xlsx')
ws = wb.active
today = date.today()
factories = []
fullname = {'MJY': '浙江美珈羽针织有限公司', 'RF': '诸暨锐锋纺织品有限公司', 'GS': '泉州罡晟轻工有限公司', 'JL': '晋江市嘉利服装织造有限公司', 'KL': '晋江市科洛服装织造有限公司', 'MQN': '泉州明全服饰有限公司', 'YYS': '湖北东林服装有限公司', 'CH': '厦门益英服装织造有限公司', 'YX': '奕鑫（厦门）皮革制品有限公司', 'RC': '诸暨市枫桥瑞城制衣厂', 'SHD': '泉州尚德服饰有限公司', 'XLY': '晋江鑫浪雅服装织造有限公司', 'SD': '泉州圣都服装有限公司', 'XKE': '泉州鑫卡尔服饰织造有限公司', 'HJ': '诸暨市汇锦服饰有限公司', 'SN-1': '桐庐逸顺服饰有限公司', 'MY': '广州市梦依服装有限公司', 'XJ': '泉州尚都服饰有限公司', 'BL': '江西省新邦服饰有限公司', 'HX': '江西万马服饰有限公司', 'HLX': '石狮市红莉祥服装织造有限公司', 'YS': '福建省源圣服饰针纺有限公司', 'SN': '杭州圣娜针纺织有限公司', 'JY': '金华市金义针织有限公司', 'CCX': '金华市骏煌服饰有限公司', 'LB': '泉州蓝博服饰有限公司', 'JH': '晋江市新塘聚汇服装有限公司', 'FY': '义乌市方圆袜业有限公司', 'MYN': '桐庐梦盈针织有限公司', 'NH': '桐庐宁航针纺有限公司', 'TZ': '泉州童真服装有限公司', 'BY': '枣庄市宝源服饰有限公司', 'HZ': '枣庄海震服饰有限公司'}
for f in ws['D']:
    if f.value is None:
        pass
    elif f.value in factories:
        pass
    elif f.value == '开票工厂':
        pass
    else:
        factories.append(f.value)
for factory in tqdm(factories, position=0, leave=True, desc=f'正在生产工厂开票登记表'):
    EDs = []
    rb = r'C:\Users\黄洁尔\PycharmProjects\test\ED+INV+REG\开票登记表\开票登记表模板.xlsx'
    rb1 = f'C:\\Users\\黄洁尔\\PycharmProjects\\test\\ED+INV+REG\\开票登记表\\{factory}.xlsx'
    shutil.copyfile(rb, rb1)
    rb1 = load_workbook(f'C:\\Users\\黄洁尔\\PycharmProjects\\test\\ED+INV+REG\\开票登记表\\{factory}.xlsx')
    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=4).value == factory and ws.cell(row=i, column=2).value not in EDs:
            EDs.append(ws.cell(row=i, column=2).value)

    rs1 = rb1.active
    rs1.title = f'{EDs[0]}'

    sheets = ['rs2', 'rs3', 'rs4', 'rs5', 'rs6', 'rs7', 'rs8', 'rs9', 'rs10', 'rs11', 'rs12', 'rs13', 'rs14',
              'rs15', 'rs16', 'rs17', 'rs18', 'rs19', 'rs20']
    for index, sheet in enumerate(sheets):
        try:
            sheet = rb1.copy_worksheet(rs1)
            sheet.title = f'{EDs[index + 1]}'
            p = 0
            for d in range(2, ws.max_row + 1):
                if ws.cell(row=d, column=8).value is None:
                    pass
                elif ws.cell(row=d, column=4).value == factory and ws.cell(row=d, column=2).value == f'{EDs[index + 1]}':
                    sheet.cell(row=1, column=2).value = fullname[ws.cell(row=d, column=4).value]
                    sheet.cell(row=2, column=2).value = '赵敏灿'
                    sheet.cell(row=7, column=4).value = '赵敏灿'
                    sheet.cell(row=7, column=7).value = f'申请日期:{today}'
                    sheet.cell(row=8, column=3).value = ws.cell(row=d, column=2).value
                    sheet.cell(row=8, column=8).value = fullname[ws.cell(row=d, column=4).value]
                    sheet.cell(row=10 + p, column=2).value = ws.cell(row=d, column=5).value
                    sheet.cell(row=10 + p, column=3).value = ws.cell(row=d, column=6).value
                    sheet.cell(row=10 + p, column=4).value = ws.cell(row=d, column=7).value
                    sheet.cell(row=10 + p, column=5).value = ws.cell(row=d, column=8).value
                    sheet.cell(row=21, column=5).value = '赵敏灿'
                    sheet.cell(row=21, column=8).value = today
                    p = p + 1
                else:
                    pass
        except IndexError:
            sheet.title = 'Sheet1'
            sheet.delete_cols(1, 11)
            break

    q = 0
    for g in range(2, ws.max_row + 1):
        if ws.cell(row=g, column=8).value is None:
            pass
        elif ws.cell(row=g, column=4).value == factory and ws.cell(row=g, column=2).value == EDs[0]:
            rs1.cell(row=1, column=2).value = fullname[ws.cell(row=g, column=4).value]
            rs1.cell(row=2, column=2).value = '赵敏灿'
            rs1.cell(row=7, column=4).value = '赵敏灿'
            rs1.cell(row=7, column=7).value = f'申请日期:{today}'
            rs1.cell(row=8, column=3).value = ws.cell(row=g, column=2).value
            rs1.cell(row=8, column=8).value = fullname[ws.cell(row=g, column=4).value]
            rs1.cell(row=10 + q, column=2).value = ws.cell(row=g, column=5).value
            rs1.cell(row=10 + q, column=3).value = ws.cell(row=g, column=6).value
            rs1.cell(row=10 + q, column=4).value = ws.cell(row=g, column=7).value
            rs1.cell(row=10 + q, column=5).value = ws.cell(row=g, column=8).value
            rs1.cell(row=21, column=5).value = '赵敏灿'
            rs1.cell(row=21, column=8).value = today
            q = q + 1
        else:
            pass

    rb1.save(f'C:\\Users\\黄洁尔\\PycharmProjects\\test\\ED+INV+REG\\开票登记表\\{factory}.xlsx')
