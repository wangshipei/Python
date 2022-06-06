import glob
import os
from openpyxl import load_workbook
from openpyxl import Workbook

wb0 = Workbook()
ws0 = wb0.active
ws0.cell(row=1, column=1).value = '指数名称'
ws0.cell(row=1, column=2).value = '指数编码'
ws0.cell(row=1, column=3).value = '登记日期'
ws0.cell(row=1, column=4).value = '指数值'
wb0.save(r'wxtest\合并数据表.xlsx')

os.chdir(r'C:\Users\黄洁尔\PycharmProjects\test\wxtest')
file_list = glob.glob('*必孚牛肉商品价格指数导入模板.xlsx')
q = 0
for file in file_list:
    wb1 = load_workbook('合并数据表.xlsx')
    ws1 = wb1.active
    wb2 = load_workbook(file)
    ws2 = wb2.active
    for i in range(2, 41):
        for j in range(1, 5):
            ws1.cell(row=int(q + i), column=j).value = ws2.cell(row=i, column=j).value
    q = q + 39
    wb1.save('合并数据表.xlsx')
    wb2.close()
print('数据合并完毕！')
