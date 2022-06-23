import re
from openpyxl import load_workbook
import shutil
import reportlab.pdfbase.ttfonts
from PyPDF2 import PdfFileReader, PdfFileWriter
import io
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib import colors
import fitz
import glob
import os

wb = load_workbook(r'C:\Users\24910\PycharmProjects\Carelabel\data\2023SS常规系列女童促销款洗水标-5.25.xlsx')
ws = wb.active

FTYs = []
for f in ws['A']:
    if f.value is None:
        pass
    elif f.value in FTYs:
        pass
    elif f.value == '工厂':
        pass
    else:
        FTYs.append(f.value)

FTY_MD_coms = []
for f1 in FTYs:
    for m in range(2, ws.max_row + 1):
        if f1 + '-' + str(ws.cell(row=m, column=9).value) + '#' + ws.cell(row=m, column=6).value in FTY_MD_coms:
            pass
        elif ws.cell(row=m, column=1).value == f1:
            FTY_MD_coms.append(f1 + '-' + str(ws.cell(row=m, column=9).value) + '#' + ws.cell(row=m, column=6).value)
        else:
            pass

arts1 = []
full_arts = []
mod_pttn = re.compile(r'\w*-(.*)#.*')
fty_pttn = re.compile(r'(#.*)')
com_pttn = re.compile(r'.*#(.*)')
tem_dir = r'C:\Users\24910\PycharmProjects\Carelabel\template'
pdf_dir = r'C:\Users\24910\PycharmProjects\Carelabel\pdf'
output_dir = r'C:\Users\24910\PycharmProjects\Carelabel\output'
for fm in FTY_MD_coms:
    for d in range(2, ws.max_row + 1):
        if ws.cell(row=d, column=1).value + '-' + str(ws.cell(row=d, column=9).value) + '#' + ws.cell(row=d,
                                                                                                      column=6).value == fm:
            full_arts.append(ws.cell(row=d, column=2).value + '---' + str(ws.cell(row=d, column=7).value) + '---' + str(
                ws.cell(row=d, column=8).value))
            arts1.append(ws.cell(row=d, column=2).value)

    arts = list(set(arts1))

    n = ''.join(re.findall(mod_pttn, fm))
    shutil.copyfile(tem_dir + '\\' + n + '.pdf', pdf_dir + '\\' + fm + '.pdf')

    pdfmetrics.registerFont(reportlab.pdfbase.ttfonts.TTFont('arial', r'C:\Windows\Fonts\arial.ttf'))
    packet = io.BytesIO()
    can = canvas.Canvas(packet)
    can.setFont('arial', 10)

    fty_model = re.sub(fty_pttn, '', fm)
    can.drawString(10, 830, fty_model)

    art_text = ' '.join(arts)
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    styleN.alignment = TA_LEFT
    styleBH = styles["Normal"]
    styleBH.alignment = TA_CENTER
    art_descrpcion = Paragraph(art_text, styleN)
    data = [[art_descrpcion]]
    art_table = Table(data)
    art_table.wrapOn(can, 50, 50)
    art_table.drawOn(can, 5, 550)

    if len(full_arts) <= 64:
        full_art_text0 = ' '.join(full_arts)
        full_art_descrpcion = Paragraph(full_art_text0, styleN)
        data = [[full_art_descrpcion]]
        full_art_table = Table(data)
        full_art_table.setStyle(
            TableStyle(
                [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black), ('BOX', (0, 0), (-1, -1), 0.25, colors.black), ]))
        full_art_table.wrapOn(can, 50, 50)
        full_art_table.drawOn(can, 150, 10)
    else:
        full_art_text1 = ' '.join(full_arts[0:65])
        full_art_descrpcion = Paragraph(full_art_text1, styleN)
        data = [[full_art_descrpcion]]
        full_art_table = Table(data)
        full_art_table.setStyle(
            TableStyle(
                [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black), ('BOX', (0, 0), (-1, -1), 0.25, colors.black), ]))
        full_art_table.wrapOn(can, 50, 50)
        full_art_table.drawOn(can, 150, 10)

        full_art_text2 = ' '.join(full_arts[64:129])
        full_art_descrpcion = Paragraph(full_art_text2, styleN)
        data = [[full_art_descrpcion]]
        full_art_table = Table(data)
        full_art_table.setStyle(
            TableStyle(
                [('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black), ('BOX', (0, 0), (-1, -1), 0.25, colors.black), ]))
        full_art_table.wrapOn(can, 50, 50)
        full_art_table.drawOn(can, 375, 10)

    composition = ''.join(re.findall(com_pttn, fm))
    can.drawString(250, 820, composition)

    can.save()
    packet.seek(0)
    new_pdf = PdfFileReader(packet)
    existing_pdf = PdfFileReader(open(pdf_dir + '\\' + fm + '.pdf', 'rb'))
    output = PdfFileWriter()
    page = existing_pdf.getPage(0)
    page.mergePage(new_pdf.getPage(0))
    output.addPage(page)
    outputStream = open(output_dir + '\\' + fm + '.pdf', 'wb')
    output.write(outputStream)
    outputStream.close()

    full_arts = []
    arts1 = []

for ft_name in FTYs:
    os.chdir(r'C:\Users\24910\PycharmProjects\Carelabel\output')
    file_list = glob.glob(f'{ft_name}*.pdf')
    result = fitz.open(file_list[0])
    for pdf in file_list[1:]:
        with fitz.open(pdf) as mfile:
            result.insert_pdf(mfile)
    result.save(output_dir + '\\' + ft_name + '.pdf')
