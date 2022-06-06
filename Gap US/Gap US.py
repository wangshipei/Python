import time
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import io
import urllib3
from urllib3.exceptions import LocationValueError


def gap_us(name, **urlplus):
    collection = list(urlplus.values())
    print(f'要爬取的所有系列为：Gap US-{name} ', '\n', collection)
    for a, a1 in urlplus.items():
        baseurl = 'https://www.gap.com/browse/category.do?cid='
        url = baseurl + str(a)
        driver = webdriver.Chrome()
        driver.get(url)
        time.sleep(10)
        driver.refresh()
        for q in range(1000, 1030):
            driver.execute_script(f'window.scrollTo(0, window.scrollY + {q});')
            time.sleep(2)
        pros = driver.find_elements(By.XPATH, '//div[@class="cat_product-image category-page-re0g4d"]/a')

        pic1s = []
        titles = []
        prices = []
        colors = []
        sizes = []
        compositions = []
        descriptions = []
        pic2s = []
        pic3s = []
        pic4s = []
        pic5s = []

        links = []
        pic1links = []
        pic2links = []
        pic3links = []
        pic4links = []
        pic5links = []

        for i in tqdm(pros, position=0, leave=True, desc=f"正在爬取Gap US-{name}-{a1}"):
            slink = i.get_attribute('href')
            try:
                driver.execute_script("window,open('');")
                driver.switch_to.window(driver.window_handles[1])
                driver.get(slink)

                try:
                    link = driver.find_element(By.XPATH, '//link[@rel="canonical"]').get_attribute('href')

                except NoSuchElementException:
                    link = ''
                except StaleElementReferenceException:
                    link = ''

                try:
                    title = driver.find_element(By.TAG_NAME, 'h1').text
                except NoSuchElementException:
                    title = ''
                except StaleElementReferenceException:
                    title = ''

                try:
                    price = driver.find_element(By.XPATH, '//div[@class="pdp-pricing pdp-mfe-hlcsuf"]').text
                except NoSuchElementException:
                    price = ''
                except StaleElementReferenceException:
                    price = ''
                try:
                    color1 = driver.find_element(By.XPATH, '//div[@class="swatch-label  pdp-mfe-194qno"]').text
                    color = re.sub('Color: ', '', color1)
                except NoSuchElementException:
                    color = ''
                except StaleElementReferenceException:
                    color = ''

                try:
                    description = driver.find_element(By.XPATH, '//meta[@name="description"]').get_attribute('content')
                except NoSuchElementException:
                    description = ''
                except StaleElementReferenceException:
                    description = ''

                try:
                    size = driver.find_element(By.XPATH,
                                               '//div[@class="pdp-mfe-17uivtt pdp-dimension pdp-dimension--auto-width"]').text
                except NoSuchElementException:
                    size = ''
                except StaleElementReferenceException:
                    size = ''

                try:
                    composition1 = driver.find_element(By.XPATH, '//script[@id="pdpData"]').get_attribute('innerHTML')
                    compttn = re.compile(r'fabric & care.*Machine wash')
                    composition2 = ''.join(re.findall(compttn, composition1))
                    composition3 = re.sub(r'fabric & care\\",\\"bulletAttributes\\":\[\\"', '', composition2)
                    composition = re.sub(r'\\",\\"Machine wash', '', composition3)
                except NoSuchElementException:
                    composition = ''
                except StaleElementReferenceException:
                    composition = ''

                try:
                    pic1link = driver.find_elements(By.XPATH, '//a[@class="hover-zoom hover-zoom-in pdp-mfe-1scitg2"]')[
                        0].get_attribute('href')
                except NoSuchElementException:
                    pic1link = ''
                except IndexError:
                    pic1link = ''
                except StaleElementReferenceException:
                    pic1link = ''

                try:
                    pic2link = driver.find_elements(By.XPATH, '//a[@class="hover-zoom hover-zoom-in pdp-mfe-1scitg2"]')[
                        1].get_attribute('href')
                except NoSuchElementException:
                    pic2link = ''
                except IndexError:
                    pic2link = ''
                except StaleElementReferenceException:
                    pic2link = ''

                try:
                    pic3link = driver.find_elements(By.XPATH, '//a[@class="hover-zoom hover-zoom-in pdp-mfe-1scitg2"]')[
                        2].get_attribute('href')
                except NoSuchElementException:
                    pic3link = ''
                except IndexError:
                    pic3link = ''
                except StaleElementReferenceException:
                    pic3link = ''

                try:
                    pic4link = driver.find_elements(By.XPATH, '//a[@class="hover-zoom hover-zoom-in pdp-mfe-1scitg2"]')[
                        3].get_attribute('href')
                except NoSuchElementException:
                    pic4link = ''
                except IndexError:
                    pic4link = ''
                except StaleElementReferenceException:
                    pic4link = ''

                try:
                    pic5link = driver.find_elements(By.XPATH, '//a[@class="hover-zoom hover-zoom-in pdp-mfe-1scitg2"]')[
                        4].get_attribute('href')
                except NoSuchElementException:
                    pic5link = ''
                except IndexError:
                    pic5link = ''
                except StaleElementReferenceException:
                    pic5link = ''

                pic1s.append('')
                titles.append(title)
                prices.append(price)
                colors.append(color)
                sizes.append(size)
                compositions.append(composition)
                descriptions.append(description)
                pic2s.append('')
                pic3s.append('')
                pic4s.append('')
                pic5s.append('')
                links.append(link)
                pic1links.append(pic1link)
                pic2links.append(pic2link)
                pic3links.append(pic3link)
                pic4links.append(pic4link)
                pic5links.append(pic5link)

                driver.close()
                driver.switch_to.window(driver.window_handles[0])

                data = pd.DataFrame()

                data['Pic-1'] = pic1s
                data['Title'] = titles
                data['Price'] = prices
                data['Color'] = colors
                data['Size'] = sizes
                data['Composition'] = compositions
                data['Description'] = descriptions
                data['Pic-2'] = pic2s
                data['Pic-3'] = pic3s
                data['Pic-4'] = pic4s
                data['Pic-5'] = pic5s
                data['Link'] = links
                data['Pic-1Link'] = pic1links
                data['Pic-2Link'] = pic2links
                data['Pic-3Link'] = pic3links
                data['Pic-4Link'] = pic4links
                data['Pic-5Link'] = pic5links

                data.to_excel(f'Gap/Gap US-{name}-{a1}.xlsx')

            except NoSuchElementException:
                print(f'此网页爬取失败：\n{slink}\n后续网页需手动处理')
                driver.close()
                driver.switch_to.window(driver.window_handles[0])

        driver.close()
        print(f'Gap US-{name}-{a1}爬取完毕！')

    print(f'Gap US-{name}所有系列爬取完毕！')


def insert_images(name, **urlplus):
    collection = '/'.join(urlplus.values())
    print('正在下载并插入图片的所有系列为Gap US-' + str(name) + '-' + str(collection))
    for a1 in urlplus.values():
        wb = load_workbook('Gap/Gap US-' + str(name) + '-' + str(a1) + '.xlsx')
        ws = wb.active

        for q1, i1 in tqdm(enumerate(ws['N'][1:], start=2), position=0, leave=True,
                           desc='正在下载并保存Gap US-' + str(name) + '-' + str(a1) + '的Pic-1'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i1.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'B{q1}')
            except LocationValueError:
                img = Image('1.jpg')
                ws.add_image(img, f'B{q1}')

        for q2, i2 in tqdm(enumerate(ws['O'][1:], start=2), position=0, leave=True,
                           desc='正在下载并保存Gap US-' + str(name) + '-' + str(a1) + '的Pic-2'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i2.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'I{q2}')
            except LocationValueError:
                img = Image('1.jpg')
                ws.add_image(img, f'I{q2}')

        for q3, i3 in tqdm(enumerate(ws['P'][1:], start=2), position=0, leave=True,
                           desc='正在下载并保存Gap US-' + str(name) + '-' + str(a1) + '的Pic-3'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i3.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'J{q3}')
            except LocationValueError:
                img = Image('1.jpg')
                ws.add_image(img, f'J{q3}')

        for q4, i4 in tqdm(enumerate(ws['Q'][1:], start=2), position=0, leave=True,
                           desc='正在下载并保存Gap US-' + str(name) + '-' + str(a1) + '的Pic-4'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i4.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'K{q4}')
            except LocationValueError:
                img = Image('1.jpg')
                ws.add_image(img, f'K{q4}')

        for q5, i5 in tqdm(enumerate(ws['R'][1:], start=2), position=0, leave=True,
                           desc='正在下载并保存Gap US-' + str(name) + '-' + str(a1) + '的Pic-5'):
            http = urllib3.PoolManager()
            try:
                r = http.request('GET', i5.value)
                image_file = io.BytesIO(r.data)
                img = Image(image_file)
                ws.add_image(img, f'L{q5}')
            except LocationValueError:
                img = Image('1.jpg')
                ws.add_image(img, f'L{q5}')

        print('Gap US-' + str(name) + '-' + str(a1) + '所有照片处理完毕')
        wb.save('Gap/Gap US-' + str(name) + '-' + str(a1) + '.xlsx')


teengirlname = 'Teen Girls'
teengirlurlplus = {
    "1159073&nav=expmore%3Ateen%3Ateen%20girls%20categories%3Atops%20%26%20t-shirts#pageId=0&department=48": "Tops & T-Shirts",
    "1159074&nav=expmore%3Ateen%3Ateen%20girls%20categories%3Asweatshirts%20%26%20outerwear#pageId=0&department=48": "Sweatshirts & Outerwear",
    "1171905&nav=expmore%3Ateen%3Ateen%20girls%20categories%3Adresses%20%26%20rompers#pageId=0&department=48": "Dresses & Rompers",
    "1172017&nav=expmore%3Ateen%3Ateen%20girls%20categories%3Aleggings%20%26%20pants#pageId=0&department=48": "Leggings & Pants",
    "1159075&nav=expmore%3Ateen%3Ateen%20girls%20categories%3Ashorts%20%26%20skirts#pageId=0&department=48": "Shorts & Skirts",
    "1159076&nav=expmore%3Ateen%3Ateen%20girls%20categories%3Ajeans#pageId=0&department=48": "Jeans",
    "1177076&nav=expmore%3Ateen%3Ateen%20girls%20categories%3Aswim#pageId=0&department=48": "Swim",
    "1177077&nav=expmore%3Ateen%3Ateen%20girls%20categories%3Aactivewear#pageId=0&department=48": "Activewear",
    "1182044&nav=expmore%3Ateen%3Ateen%20girls%20categories%3Apajamas#pageId=0&department=48": "Pajamas"}

teenboyname = 'Teen Boys'
teenboyurlplus = {
    "1161314&nav=expmore%3Ateen%3Ateen%20guys%20categories%3Atops%20%26%20t-shirts#pageId=0&department=16": "Tops & T-Shirts",
    "1161311&nav=expmore%3Ateen%3Ateen%20guys%20categories%3Asweatshirts%20%26%20outerwear#pageId=0&department=16": "Sweatshirts & Outerwear",
    "1172009&nav=expmore%3Ateen%3Ateen%20guys%20categories%3Apants%20%26%20joggers#pageId=0&department=16": "Pants & Joggers",
    "1171909&nav=expmore%3Ateen%3Ateen%20guys%20categories%3Ashorts#pageId=0&department=16": "Shorts",
    "1161313&nav=expmore%3Ateen%3Ateen%20guys%20categories%3Ajeans#pageId=0&department=16": "Jeans",
    "1177079&nav=expmore%3Ateen%3Ateen%20guys%20categories%3Aactivewear#pageId=0&department=16": "Activewear",
    "1184747&nav=expmore%3Ateen%3Ateen%20guys%20categories%3Apajamas#pageId=0&department=16": "Pajamas"
}

girlname = 'Girls'
girlurlplus = {
    "6276&nav=expmore%3Agapkids%3Acategories%3Ajeans#pageId=0&department=48": "Jeans",
    "6300&nav=expmore%3Agapkids%3Acategories%3Adresses#pageId=0&department=48": "Dresses",
    "1141739&nav=expmore%3Agapkids%3Acategories%3Arompers%20%26%20jumpsuits#pageId=0&department=48": "Rompers & Jumpsuits",
    "14417&nav=expmore%3Agapkids%3Acategories%3Atops%20%26%20t-shirts#pageId=0&department=48": "Tops & T-Shirts",
    "1122942&nav=expmore%3Agapkids%3Acategories%3Agraphic%20t-shirts#pageId=0&department=48": "Graphic T-Shirts",
    "1153699&nav=expmore%3Agapkids%3Acategories%3Ashirts%20%26%20polos#pageId=0&department=48": "Shirts & Polos",
    "1056270&nav=expmore%3Agapkids%3Acategories%3Asweatshirts%20%26%20sweatpants#pageId=0&department=48": "Sweatshirts & Sweatpants",
    "1161843&nav=expmore%3Agapkids%3Acategories%3Asweaters#pageId=0&department=48": "Sweaters",
    "6303&nav=expmore%3Agapkids%3Acategories%3Aouterwear%20%26%20jackets#pageId=0&department=48": "Outerwear & Jackets",
    "13148&nav=expmore%3Agapkids%3Acategories%3Aleggings%20%26%20pants#pageId=0&department=48": "Leggings & Pants",
    "14403&nav=expmore%3Agapkids%3Acategories%3Ashorts%20%26%20skirts#pageId=0&department=48": "Shorts & Skirts",
    "1051487&nav=expmore%3Agapkids%3Acategories%3Agapfit%20%26%20active#pageId=0&department=48": "GapFit & Active",
    "1075777&nav=expmore%3Agapkids%3Acategories%3Aswim#pageId=0&department=48": "Swim",
    "6323&nav=expmore%3Agapkids%3Acategories%3Apajamas#pageId=0&department=48": "Pajamas",
    "1061822&nav=expmore%3Agapkids%3Acategories%3Aschool%20uniforms#pageId=0&department=48": "School Uniforms",
    "1188764&nav=expmore%3Agapkids%3Acategories%3Amulti-packs#pageId=0&department=48": "Multi-Packs",
    "1189408&nav=expmore%3Agapkids%3Acategories%3Amatching%20sets#pageId=0&department=48": "Matching Sets",
    "1107336&nav=expmore%3Agapkids%3Acategories%3Asocks%20%26%20underwear#pageId=0&department=48": "Socks & Underwear",
    "56233&nav=expmore%3Agapkids%3Acategories%3Aaccessories%20%26%20more#pageId=0&department=48": "Accessories & More"
}

boyname = 'Boys'
boyurlplus = {
    "6189&nav=expmore%3Agapkids%3Acategories%3Ajeans#pageId=0&department=16": "Jeans",
    "1122119&nav=expmore%3Agapkids%3Acategories%3Agraphic%20t-shirts#pageId=0&department=16": "Graphic T-Shirts",
    "1070923&nav=expmore%3Agapkids%3Acategories%3At-shirts#pageId=0&department=16": "T-Shirts",
    "6197&nav=expmore%3Agapkids%3Acategories%3Ashirts%20%26%20polos#pageId=0&department=16": "Shirts & Polos",
    "1117991&nav=expmore%3Agapkids%3Acategories%3Asweatshirts%20%26%20sweatpants#pageId=0&department=16": "Sweatshirts & Sweatpants",
    "1175613&nav=expmore%3Agapkids%3Acategories%3Asweaters#pageId=0&department=16": "Sweaters",
    "6205&nav=expmore%3Agapkids%3Acategories%3Aouterwear%20%26%20jackets#pageId=0&department=16": "Outerwear & Jackets",
    "1085428&nav=expmore%3Agapkids%3Acategories%3Ajoggers%20%26%20sweatpants#pageId=0&department=16": "Joggers & Sweatpants",
    "6187&nav=expmore%3Agapkids%3Acategories%3Apants#pageId=0&department=16": "Pants",
    "6191&nav=expmore%3Agapkids%3Acategories%3Ashorts#pageId=0&department=16": "Shorts",
    "1050851&nav=expmore%3Agapkids%3Acategories%3Agapfit%20%26%20active#pageId=0&department=16": "GapFit & Active",
    "1075793&nav=expmore%3Agapkids%3Acategories%3Aswim#pageId=0&department=16": "Swim",
    "9470&nav=expmore%3Agapkids%3Acategories%3Apajamas#pageId=0&department=16": "Pajamas",
    "1060990&nav=expmore%3Agapkids%3Acategories%3Aschool%20uniforms#pageId=0&department=16": "School Uniforms",
    "1188761&nav=expmore%3Agapkids%3Acategories%3Amulti-packs#pageId=0&department=16": "Multi-Packs",
    "1107335&nav=expmore%3Agapkids%3Acategories%3Asocks%20%26%20underwear#pageId=0&department=16": "Socks & Underwear",
    "96875&nav=expmore%3Agapkids%3Acategories%3Aaccessories%20%26%20more#pageId=0&department=16": "Accessories & More"
}

toddlername = 'Toddler'
toddlerurlplus = {
    "6427&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3Ajeans#pageId=0&department=165": "Jeans",
    "6436&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3Adresses#pageId=0&department=165": "Dresses",
    "1145378&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3Arompers%20%26%20jumpsuits#pageId=0&department=165": "Rompers & Jumpsuits",
    "6444&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3At-shirts%20%26%20graphics#pageId=0&department=165": "T-Shirts & Graphics",
    "1132758&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3Atops#pageId=0&department=165": "Tops",
    "17846&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3Asweatshirts%20%26%20sweatpants#pageId=0&department=165": "Sweatshirts & Sweatpants",
    "1175977&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3Asweaters#pageId=0&department=165": "Sweaters",
    "8770&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3Aouterwear%20%26%20jackets#pageId=0&department=165": "Outerwear & Jackets",
    "12378&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3Aleggings%20%26%20pants#pageId=0&department=165": "Leggings & Pants",
    "1121815&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3Ashorts%20%26%20skirts#pageId=0&department=165": "Shorts & Skirts",
    "76918&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3Apajamas#pageId=0&department=165": "Pajamas",
    "1072981&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3Aswim#pageId=0&department=165": "Swim",
    "1188469&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3Amulti-packs#pageId=0&department=165": "Multi-Packs",
    "1084375&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3Aaccessories%20%26%20more#pageId=0&department=165": "Accessories & More",
    "1033898&nav=expmore%3Ababygap%3Atoddler%20girl%2012m%20to%205y%3Asocks%20%26%20underwear#pageId=0&department=165": "Socks & Underwear",
    "6359&nav=expmore%3Ababygap%3Atoddler%20boy%2012m%20to%205y%3Ajeans#pageId=0&department=166": "Jeans",
    "1016096&nav=expmore%3Ababygap%3Atoddler%20boy%2012m%20to%205y%3At-shirts%20%26%20graphics#pageId=0&department=166": "T-Shirts & Graphics",
    "1016169&nav=expmore%3Ababygap%3Atoddler%20boy%2012m%20to%205y%3Ashirts%20%26%20polos#pageId=0&department=166": "Shirts & Polos",
    "1016107&nav=expmore%3Ababygap%3Atoddler%20boy%2012m%20to%205y%3Asweatshirts%20%26%20sweatpants#pageId=0&department=166": "Sweatshirts & Sweatpants",
    "1175972&nav=expmore%3Ababygap%3Atoddler%20boy%2012m%20to%205y%3Asweaters#pageId=0&department=166": "Sweaters",
    "1016108&nav=expmore%3Ababygap%3Atoddler%20boy%2012m%20to%205y%3Aouterwear%20%26%20jackets#pageId=0&department=166": "Outerwear & Jackets",
    "1016106&nav=expmore%3Ababygap%3Atoddler%20boy%2012m%20to%205y%3Apants#pageId=0&department=166": "Pants",
    "1121839&nav=expmore%3Ababygap%3Atoddler%20boy%2012m%20to%205y%3Ashorts#pageId=0&department=166": "Shorts",
    "1016109&nav=expmore%3Ababygap%3Atoddler%20boy%2012m%20to%205y%3Apajamas#pageId=0&department=166": "Pajamas",
    "1072982&nav=expmore%3Ababygap%3Atoddler%20boy%2012m%20to%205y%3Aswim#pageId=0&department=166": "Swim",
    "1188470&nav=expmore%3Ababygap%3Atoddler%20boy%2012m%20to%205y%3Amulti-packs#pageId=0&department=166": "Multi-Packs",
    "1084376&nav=expmore%3Ababygap%3Atoddler%20boy%2012m%20to%205y%3Aaccessories%20%26%20more#pageId=0&department=166": "Accessories & More",
    "1034157&nav=expmore%3Ababygap%3Atoddler%20boy%2012m%20to%205y%3Asocks%20%26%20underwear#pageId=0&department=166": "Socks & Underwear",

}

babyname = 'Baby'
babyurlplus = {
    "1098333&nav=expmore%3Ababygap%3Ababy%20girl%200%20to%2024m%3Asets#pageId=0&department=165": "Sets",
    "1027203&nav=expmore%3Ababygap%3Ababy%20girl%200%20to%2024m%3Aone-pieces#pageId=0&department=165": "One-pieces",
    "6437&nav=expmore%3Ababygap%3Ababy%20girl%200%20to%2024m%3Adresses%20%26%20rompers#pageId=0&department=165": "Dresses & Rompers",
    "7189&nav=expmore%3Ababygap%3Ababy%20girl%200%20to%2024m%3Abodysuits%20%26%20tops#pageId=0&department=165": "Bodysuits & Tops",
    "1028587&nav=expmore%3Ababygap%3Ababy%20girl%200%20to%2024m%3Asweatshirts%20%26%20sweatpants#pageId=0&department=165": "Sweatshirts & Sweatpants",
    "1175810&nav=expmore%3Ababygap%3Ababy%20girl%200%20to%2024m%3Asweaters#pageId=0&department=165": "Sweaters",
    "1108608&nav=expmore%3Ababygap%3Ababy%20girl%200%20to%2024m%3Aouterwear%20%26%20jackets#pageId=0&department=165": "Outerwear & Jackets",
    "7191&nav=expmore%3Ababygap%3Ababy%20girl%200%20to%2024m%3Ajeans%2C%20pants%20%26%20leggings#pageId=0&department=165": "Jeans, Pants & Leggings",
    "1102200&nav=expmore%3Ababygap%3Ababy%20girl%200%20to%2024m%3Ashorts%20%26%20skirts#pageId=0&department=165": "Shorts & Skirts",
    "76748&nav=expmore%3Ababygap%3Ababy%20girl%200%20to%2024m%3Apajamas#pageId=0&department=165": "Pajamas",
    "1120931&nav=expmore%3Ababygap%3Ababy%20girl%200%20to%2024m%3Aswim#pageId=0&department=165": "Swim",
    "1146519&nav=expmore%3Ababygap%3Ababy%20girl%200%20to%2024m%3Amulti-packs#pageId=0&department=165": "Multi-Packs",
    "1086661&nav=expmore%3Ababygap%3Ababy%20girl%200%20to%2024m%3Aaccessories%20%26%20more#pageId=0&department=165": "Accessories & More",
    "1098335&nav=expmore%3Ababygap%3Ababy%20boy%200%20to%2024m%3Asets#pageId=0&department=166": "Sets",
    "1027202&nav=expmore%3Ababygap%3Ababy%20boy%200%20to%2024m%3Aone-pieces#pageId=0&department=166": "One-pieces",
    "95598&nav=expmore%3Ababygap%3Ababy%20boy%200%20to%2024m%3Abodysuits%20%26%20tops#pageId=0&department=166": "Bodysuits & Tops",
    "1028588&nav=expmore%3Ababygap%3Ababy%20boy%200%20to%2024m%3Asweatshirts%20%26%20sweatpants#pageId=0&department=166": "Sweatshirts & Sweatpants",
    "1175813&nav=expmore%3Ababygap%3Ababy%20boy%200%20to%2024m%3Asweaters#pageId=0&department=166": "Sweaters",
    "1114529&nav=expmore%3Ababygap%3Ababy%20boy%200%20to%2024m%3Aouterwear%20%26%20jackets#pageId=0&department=166": "Outerwear & Jackets",
    "95684&nav=expmore%3Ababygap%3Ababy%20boy%200%20to%2024m%3Ajeans%20%26%20pants#pageId=0&department=166": "Jeans & Pants",
    "1102201&nav=expmore%3Ababygap%3Ababy%20boy%200%20to%2024m%3Ashorts#pageId=0&department=166": "Shorts",
    "95697&nav=expmore%3Ababygap%3Ababy%20boy%200%20to%2024m%3Apajamas#pageId=0&department=166": "Pajamas",
    "1120932&nav=expmore%3Ababygap%3Ababy%20boy%200%20to%2024m%3Aswim#pageId=0&department=165": "Swim",
    "1146528&nav=expmore%3Ababygap%3Ababy%20boy%200%20to%2024m%3Amulti-packs#pageId=0&department=166": "Multi-Packs",
    "1086662&nav=expmore%3Ababygap%3Ababy%20boy%200%20to%2024m%3Aaccessories%20%26%20more#pageId=0&department=166": "Accessories & More",
}

gap_us(teengirlname, **teengirlurlplus)
gap_us(teenboyname, **teenboyurlplus)
gap_us(girlname, **girlurlplus)
gap_us(boyname, **boyurlplus)
gap_us(toddlername, **toddlerurlplus)
gap_us(babyname, **babyurlplus)

insert_images(teengirlname, **teengirlurlplus)
insert_images(teenboyname, **teenboyurlplus)
insert_images(girlname, **girlurlplus)
insert_images(boyname, **boyurlplus)
insert_images(toddlername, **toddlerurlplus)
insert_images(babyname, **babyurlplus)
